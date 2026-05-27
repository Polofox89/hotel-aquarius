#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tagesbuffet – Core-Logik (ohne GUI / ohne tkinter)
====================================================

Funktionen, die sowohl von der Desktop-GUI (tagesbuffet_gui.py) als auch
vom Web-Backend (web_app.py) verwendet werden:

- KI-Kategorisierung via Claude API
- Slot-Konfiguration für 11 Eingabefelder
- Default-Vorschläge pro Kategorie
- Aufbau der Dropdown-Vorschläge aus dem Excel-Archiv

Dieses Modul lädt KEIN tkinter und keine GUI-Pakete, daher kann es
auch auf einem headless Server (Linux ohne Display) importiert werden.
"""

from __future__ import annotations

import collections
import json
from pathlib import Path

from tagesbuffet_generator import CATEGORIES


# ── Optionale Abhängigkeiten (gracefully handled) ─────────────────────────────

try:
    from openpyxl import load_workbook
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    load_workbook = None  # type: ignore

try:
    import anthropic
    ANTHROPIC_OK = True
except ImportError:
    ANTHROPIC_OK = False
    anthropic = None  # type: ignore


# ── Slot-Konfiguration ────────────────────────────────────────────────────────
#
# Jeder Eintrag: (Combobox-Key, Kategorie-Key, Slot-Index, Label-Key, optional)
#
SLOTS = [
    ("suppe_1",      "suppe",       0, "cat_suppe",       False),
    ("haupt_1",      "haupt",       0, "cat_haupt",       False),
    ("haupt_2",      "haupt",       1, "cat_haupt",       False),
    ("pasta_1",      "pasta",       0, "cat_pasta",       True),
    ("beilage_1",    "beilagen",    0, "cat_beilage",     False),
    ("beilage_2",    "beilagen",    1, "cat_beilage",     False),
    ("beilage_3",    "beilagen",    2, "cat_beilage",     False),
    ("beilage_4",    "beilagen",    3, "cat_beilage",     True),
    ("salate_1",     "salate",      0, "cat_salate",      False),
    ("partypfanne_1","partypfanne", 0, "cat_partypfanne", False),
    ("dessert_1",    "dessert",     0, "cat_dessert",     False),
]

# Mapping Excel-Label (Großbuchstaben) → Kategorie-Key
LABEL_TO_KEY = {label: key for key, label, *_ in CATEGORIES}


# ── Default-Vorschläge pro Kategorie ──────────────────────────────────────────

DEFAULT_SUGGESTIONS = {
    "suppe": [
        "Tomatensuppe", "Hühnerbrühe mit Nudeln", "Kürbissuppe", "Linsensuppe",
        "Spargelcremesuppe", "Gulaschsuppe", "Kartoffelsuppe", "Erbsensuppe",
        "Brokkolicremesuppe", "Tagessuppe",
    ],
    "haupt": [
        "Wiener Schnitzel", "Kasseler", "Hähnchenbrust", "Rinderbraten",
        "Putenrollbraten", "Schweinebraten", "Frikadellen", "Lachsfilet",
        "Hähnchenschenkel", "Rouladen",
    ],
    "pasta": [
        "Spaghetti Bolognese", "Tortellini in Sahnesauce", "Penne Arrabbiata",
        "Lasagne", "Spaghetti Carbonara", "Tagliatelle mit Lachs",
        "Nudelauflauf", "Rigatoni mit Tomatensauce", "Cannelloni",
        "Spätzle mit Käse",
    ],
    "beilagen": [
        "Bratkartoffeln", "Salzkartoffeln", "Kartoffelpüree", "Reis",
        "Spätzle", "Pommes frites", "Brokkoli", "Erbsen und Möhren",
        "Blumenkohl", "Rotkohl",
    ],
    "salate": [
        "Salatbuffet", "Gemischter Salat", "Tomatensalat", "Gurkensalat",
        "Krautsalat", "Kartoffelsalat", "Nudelsalat", "Heringssalat",
        "Bohnensalat", "Rote-Bete-Salat",
    ],
    "partypfanne": [
        "Tagesempfehlung", "Hähnchen Süß-Sauer", "Chili con Carne",
        "Bauernpfanne", "Gyrospfanne", "Geschnetzeltes Züricher Art",
        "Currywurst-Pfanne", "Nudelpfanne", "Reispfanne", "Gemüsepfanne",
    ],
    "dessert": [
        "Dessertbuffet", "Vanillepudding", "Schokopudding", "Obstsalat",
        "Mousse au Chocolat", "Rote Grütze", "Tiramisu", "Apfelstrudel",
        "Eis mit heißen Kirschen", "Quarkspeise",
    ],
}


# ── Vorschläge aus Excel-Archiv ───────────────────────────────────────────────

def build_suggestions_from_excel(archiv_file: Path, top_n: int | None = None) -> dict:
    """
    Liest das Excel-Archiv und gibt ALLE jemals verwendeten Gerichte pro
    Kategorie zurück. Sortierung: alphabetisch (case-insensitive).
    DEFAULT_SUGGESTIONS werden als Sockel hinzugefügt, falls fehlend.
    """
    base_defaults = {k: list(v) for k, v in DEFAULT_SUGGESTIONS.items()}
    for key, *_ in CATEGORIES:
        base_defaults.setdefault(key, [])

    if not (OPENPYXL_OK and archiv_file.exists()):
        result = {
            k: sorted(set(v), key=str.lower)
            for k, v in base_defaults.items()
        }
        return {k: (v[:top_n] if top_n else v) for k, v in result.items()}

    try:
        wb = load_workbook(archiv_file, read_only=True, data_only=True)
        ws = wb.active
        used: dict = {key: set() for key, *_ in CATEGORIES}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            kategorie, speise = row[1], row[2]
            if not kategorie or not speise:
                continue
            key = LABEL_TO_KEY.get(str(kategorie).strip().upper())
            if key:
                used[key].add(str(speise).strip())
        wb.close()

        result: dict = {}
        for key, *_ in CATEGORIES:
            names = set(used[key])
            names.update(base_defaults.get(key, []))   # Defaults als Sockel
            sorted_names = sorted(names, key=str.lower)
            result[key] = sorted_names[:top_n] if top_n else sorted_names
        return result
    except Exception as e:
        print(f"Warnung: Vorschläge aus Excel konnten nicht gelesen werden – {e}")
        result = {
            k: sorted(set(v), key=str.lower)
            for k, v in base_defaults.items()
        }
        return {k: (v[:top_n] if top_n else v) for k, v in result.items()}


# ── KI-Kategorisierung ────────────────────────────────────────────────────────

SYSTEM_PROMPT = """Du bist Assistent für das Hotel Aquarius in Norddeich.
Analysiere den eingegebenen Text und ordne die Speisen den Buffet-Kategorien zu.
Der Eingabetext kann auf Deutsch oder Arabisch sein – die Ausgabe ist IMMER Deutsch.

Gib NUR gültiges JSON zurück – kein Markdown, keine Erklärungen.

JSON-Schema:
{
  "suppe":       ["..."],
  "haupt":       ["...", "..."],
  "pasta":       ["..."],
  "beilagen":    ["...", "..."],
  "salate":      ["Salatbuffet"],
  "partypfanne": ["..."],
  "dessert":     ["Dessertbuffet"]
}

Regeln:
- Schreibe Speisen IMMER auf Deutsch und mit großem Anfangsbuchstaben
- Wenn eine Kategorie fehlt, nutze sinnvolle Standardwerte
- Salate → immer "Salatbuffet" (außer explizit anders angegeben)
- Dessert → immer "Dessertbuffet" (außer explizit anders angegeben)
- Partypfanne → wenn nicht genannt, schreibe "Tagesempfehlung"
- Pasta → EIGENE Kategorie für alle Nudelgerichte (Spaghetti, Tortellini,
  Lasagne, Penne, Tagliatelle, Cannelloni, Rigatoni, Spätzle mit Käse, ...).
  Nudelgerichte gehören NICHT zu "haupt", sondern in "pasta".
  Wenn keine Pasta genannt ist, gib eine leere Liste zurück: "pasta": []
- "haupt" enthält NUR Fleisch/Fisch/Geflügel-Hauptgerichte, KEINE Nudelgerichte
- Beilagen: Gemüse zuerst, dann Stärke
- Gib NUR das JSON zurück, sonst nichts"""


def ki_kategorisieren(text: str, api_key: str) -> dict:
    """Sendet Freitext an Claude und gibt das kategorisierte Menü zurück."""
    if not ANTHROPIC_OK:
        raise RuntimeError("anthropic-Paket nicht installiert (pip install anthropic).")

    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=512,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": text}],
    )
    raw = message.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.strip()
    menu = json.loads(raw)

    defaults = {
        "suppe":       [],
        "haupt":       [],
        "pasta":       [],
        "beilagen":    [],
        "salate":      ["Salatbuffet"],
        "partypfanne": ["Tagesempfehlung"],
        "dessert":     ["Dessertbuffet"],
    }
    for key, default in defaults.items():
        if key not in menu:
            menu[key] = default
        elif not menu[key] and default:
            menu[key] = default
    return menu
