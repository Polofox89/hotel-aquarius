#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tagesbuffet-GUI – Hotel Aquarius
==================================
Tkinter-Oberfläche für den Tagesbuffet-Generator.

Features:
 - Dropdown-Comboboxen mit Top-10-Vorschlägen aus dem Excel-Archiv
   (zusätzlich Freitext für nicht gelistete Gerichte)
 - Live-Vorschau: Bild aktualisiert sich bei jeder Änderung
 - Sprachumschalter Deutsch / Arabisch (Bildausgabe bleibt deutsch)
 - Konfigurierbare Pfade für Bilder (z. B. Google-Drive-Sync)
   und Excel-Archiv (lokal)
 - KI-Freitext: Speisen können auch frei eingegeben und automatisch
   in die Kategorien einsortiert werden

Voraussetzungen (einmalig installieren):
    pip install pillow openpyxl anthropic tkcalendar

Verwendung:
    python tagesbuffet_gui.py
"""

import os
import sys
import json
import threading
import collections
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

# ── Pillow ────────────────────────────────────────────────────────────────────
try:
    from PIL import Image, ImageTk
    PIL_OK = True
except ImportError:
    PIL_OK = False
    print("Warnung: Pillow nicht installiert – Vorschau deaktiviert.")
    print("         pip install Pillow")

# ── openpyxl (für Top-10-Vorschläge aus dem Archiv) ───────────────────────────
try:
    from openpyxl import load_workbook
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Claude API ────────────────────────────────────────────────────────────────
try:
    import anthropic
    ANTHROPIC_OK = True
except ImportError:
    ANTHROPIC_OK = False
    print("Warnung: anthropic nicht installiert – KI-Erkennung deaktiviert.")
    print("         pip install anthropic")

# ── Generator importieren ─────────────────────────────────────────────────────
try:
    sys.path.insert(0, str(Path(__file__).parent))
    from tagesbuffet_generator import (
        TagesbuffetGenerator, CATEGORIES, LOGO_PATH,
    )
    GENERATOR_OK = True
except ImportError as e:
    GENERATOR_OK = False
    print(f"Fehler: tagesbuffet_generator.py nicht gefunden – {e}")


# ══════════════════════════════════════════════════════════════════════════════
# Farben & Stil  (Aquarius Teal)
# ══════════════════════════════════════════════════════════════════════════════

C_TEAL       = "#4BBFBF"
C_TEAL_DARK  = "#1A8C96"
C_TEAL_LIGHT = "#e0f5f5"
C_BG         = "#f7fafa"
C_WHITE      = "#ffffff"
C_TEXT       = "#2c2c2c"
C_GRAY       = "#888888"
C_BORDER     = "#cccccc"
C_SUCCESS    = "#2e7d32"
C_ERROR      = "#c62828"

FONT_TITLE   = ("Segoe UI", 16, "bold")
FONT_LABEL   = ("Segoe UI", 10, "bold")
FONT_NORMAL  = ("Segoe UI", 10)
FONT_SMALL   = ("Segoe UI", 9)
FONT_TINY    = ("Segoe UI", 8)

# Eingabe-Spalte: ca. 30 % größer (für bessere Lesbarkeit)
FONT_IN_FRAME  = ("Segoe UI", 13, "bold")   # LabelFrame-Titel
FONT_IN_LABEL  = ("Segoe UI", 13, "bold")   # Kategorie-Labels
FONT_IN_NORMAL = ("Segoe UI", 13)           # Eingabefelder, Comboboxen
FONT_IN_SMALL  = ("Segoe UI", 12)           # Hinweise / API-Key
FONT_IN_DATE   = ("Segoe UI", 16, "bold")   # Datum (groß, war 12)
FONT_IN_BTN    = ("Segoe UI", 12)           # Datums-Buttons
FONT_IN_KI     = ("Segoe UI", 13, "bold")   # KI-Analyse-Button


# ══════════════════════════════════════════════════════════════════════════════
# Sprachen / i18n
# ══════════════════════════════════════════════════════════════════════════════

TRANSLATIONS = {
    "de": {
        "app_title":           "Tagesbuffet-Generator – Hotel Aquarius",
        "header_title":        "TAGESBUFFET-GENERATOR",
        "header_subtitle":     "Hotel Aquarius · Norddeich",
        "btn_path_images":     "📁 Bilder",
        "btn_path_excel":      "📊 Excel",
        "lang_switch":         "DE | AR",

        "frame_date":          " 📅 Datum ",
        "btn_today":           "Heute",
        "btn_tomorrow":        "Morgen",
        "btn_calendar":        "📅 Kalender",

        "frame_freitext":      " ✏ Speisen frei eingeben (KI ordnet zu) ",
        "hint_freitext":       "Einfach alle Speisen eingeben – kommagetrennt oder zeilenweise:",
        "placeholder_freitext":"Beispiel:\nTomatensuppe, Wiener Schnitzel, Kasseler, "
                                "Brokkoli, Bratkartoffeln, Reis, Tagessuppe",
        "btn_ki":              "🤖  KI-Analyse  →  Felder automatisch füllen",
        "btn_ki_running":      "⏳  KI analysiert ...",

        "frame_menu":          " 📋 Menü zusammenstellen ",
        "cat_suppe":           "🥣 Suppe",
        "cat_haupt":           "🍖 Hauptgericht {n}",
        "cat_pasta":           "🍝 Pasta",
        "cat_beilage":         "🥦 Beilage {n}",
        "cat_salate":          "🥗 Salate",
        "cat_partypfanne":     "🍳 Partypfanne",
        "cat_dessert":         "🍮 Dessert",
        "hint_optional":       "(optional)",

        "frame_apikey":        " 🔑 API-Key (Anthropic) ",
        "lbl_apikey_hint":     "ANTHROPIC_API_KEY – wird nur lokal gespeichert",

        "frame_preview":       " 👁 Live-Vorschau ",
        "btn_save_image":      "🖼  Bild speichern",
        "btn_save_excel":      "📊  Excel-Archiv",
        "btn_full_preview":    "👁  Vollansicht",
        "frame_history":       " 🕐 Letzte 7 Tage ",
        "history_hint":        "Klicken zum Übernehmen",
        "history_empty":       "Noch keine Einträge.",

        "status_ready":        "Bereit.",
        "status_rendering":    "Vorschau wird aktualisiert ...",
        "status_ki_running":   "KI analysiert die Eingabe ...",
        "status_ki_ok":        "✓ Kategorien erkannt.",
        "status_image_saved":  "✓ Bild gespeichert: {name}",
        "status_excel_saved":  "✓ Excel-Archiv aktualisiert.",
        "status_path_changed": "✓ Speicherort geändert.",

        "msg_no_input_title":     "Eingabe fehlt",
        "msg_no_input":           "Bitte zuerst Speisen eingeben.",
        "msg_no_apikey_title":    "API-Key fehlt",
        "msg_no_apikey":          "Bitte den Anthropic API-Key eingeben oder als\n"
                                  "Umgebungsvariable ANTHROPIC_API_KEY setzen.",
        "msg_no_anthropic_title": "API nicht verfügbar",
        "msg_no_anthropic":       "anthropic-Paket nicht installiert.\n\npip install anthropic",
        "msg_ki_error_title":     "KI-Fehler",
        "msg_ki_error":           "Analyse fehlgeschlagen:\n{err}",
        "msg_menu_empty_title":   "Menü leer",
        "msg_menu_empty":         "Bitte erst Speisen eingeben.",
        "msg_image_error_title":  "Bild-Fehler",
        "msg_image_error":        "Bild konnte nicht erstellt werden:\n{err}",
        "msg_excel_error_title":  "Excel-Fehler",
        "msg_no_image_title":     "Kein Bild",
        "msg_no_image":           "Bitte zuerst ein Bild erstellen.",
        "msg_dialog_image_title": "Speicherort für Bilder wählen (z. B. Google-Drive-Ordner)",
        "msg_dialog_excel_title": "Speicherort für Excel-Archiv wählen (lokal)",
        "msg_open_in_explorer":   "Im Explorer öffnen",
    },
    "ar": {
        "app_title":           "مولّد بوفيه اليوم – فندق أكواريوس",
        "header_title":        "مولّد بوفيه اليوم",
        "header_subtitle":     "فندق أكواريوس · نوردايش",
        "btn_path_images":     "📁 الصور",
        "btn_path_excel":      "📊 إكسل",
        "lang_switch":         "AR | DE",

        "frame_date":          " 📅 التاريخ ",
        "btn_today":           "اليوم",
        "btn_tomorrow":        "غداً",
        "btn_calendar":        "📅 التقويم",

        "frame_freitext":      " ✏ إدخال الأطباق (الذكاء الاصطناعي يصنّفها) ",
        "hint_freitext":       "أدخل جميع الأطباق – مفصولة بفاصلة أو في أسطر:",
        "placeholder_freitext":"مثال:\nشوربة الطماطم، شنيتزل فيينا، كاسيلر، "
                                "بروكلي، بطاطس مقلية، أرز، شوربة اليوم",
        "btn_ki":              "🤖  تحليل بالذكاء الاصطناعي  →  تعبئة الحقول",
        "btn_ki_running":      "⏳  جارٍ التحليل ...",

        "frame_menu":          " 📋 تجميع القائمة ",
        "cat_suppe":           "🥣 الحساء",
        "cat_haupt":           "🍖 الطبق الرئيسي {n}",
        "cat_pasta":           "🍝 معكرونة",
        "cat_beilage":         "🥦 الطبق الجانبي {n}",
        "cat_salate":          "🥗 السلطات",
        "cat_partypfanne":     "🍳 مقلاة الحفلة",
        "cat_dessert":         "🍮 الحلوى",
        "hint_optional":       "(اختياري)",

        "frame_apikey":        " 🔑 مفتاح API (Anthropic) ",
        "lbl_apikey_hint":     "ANTHROPIC_API_KEY – يُحفظ محلياً فقط",

        "frame_preview":       " 👁 المعاينة المباشرة ",
        "btn_save_image":      "🖼  حفظ الصورة",
        "btn_save_excel":      "📊  أرشيف Excel",
        "btn_full_preview":    "👁  عرض كامل",
        "frame_history":       " 🕐 آخر 7 أيام ",
        "history_hint":        "اضغط للاستعادة",
        "history_empty":       "لا توجد سجلات بعد.",

        "status_ready":        "جاهز.",
        "status_rendering":    "جارٍ تحديث المعاينة ...",
        "status_ki_running":   "الذكاء الاصطناعي يحلّل ...",
        "status_ki_ok":        "✓ تم التعرّف على الفئات.",
        "status_image_saved":  "✓ تم حفظ الصورة: {name}",
        "status_excel_saved":  "✓ تم تحديث أرشيف Excel.",
        "status_path_changed": "✓ تم تغيير مكان الحفظ.",

        "msg_no_input_title":     "الإدخال مفقود",
        "msg_no_input":           "يرجى إدخال الأطباق أولاً.",
        "msg_no_apikey_title":    "مفتاح API مفقود",
        "msg_no_apikey":          "يرجى إدخال مفتاح Anthropic API\n"
                                  "أو ضبطه عبر متغير البيئة ANTHROPIC_API_KEY.",
        "msg_no_anthropic_title": "API غير متاح",
        "msg_no_anthropic":       "حزمة anthropic غير مُثبّتة.\n\npip install anthropic",
        "msg_ki_error_title":     "خطأ في الذكاء الاصطناعي",
        "msg_ki_error":           "فشل التحليل:\n{err}",
        "msg_menu_empty_title":   "القائمة فارغة",
        "msg_menu_empty":         "يرجى إدخال الأطباق أولاً.",
        "msg_image_error_title":  "خطأ في الصورة",
        "msg_image_error":        "تعذّر إنشاء الصورة:\n{err}",
        "msg_excel_error_title":  "خطأ في Excel",
        "msg_no_image_title":     "لا توجد صورة",
        "msg_no_image":           "يرجى إنشاء صورة أولاً.",
        "msg_dialog_image_title": "اختيار مكان حفظ الصور (مثلاً مجلد Google Drive)",
        "msg_dialog_excel_title": "اختيار مكان حفظ أرشيف Excel (محلي)",
        "msg_open_in_explorer":   "فتح في المستكشف",
    },
}


# ══════════════════════════════════════════════════════════════════════════════
# KI-Kategorisierung via Claude API
# ══════════════════════════════════════════════════════════════════════════════

SYSTEM_PROMPT = """Du bist Assistent für das Hotel Aquarius in Norddeich.
Analysiere den eingegebenen Text und ordne die Speisen den Buffet-Kategorien zu.
Der Eingabetext kann auf Deutsch oder Arabisch sein – die Ausgabe ist IMMER Deutsch.

Gib NUR gültiges JSON zurück – kein Markdown, keine Erklärungen.

JSON-Schema:
{
  "suppe":       ["..."],
  "haupt":       ["...", "..."],
  "pasta":       ["..."],
  "beilagen":    ["...", "..."],
  "salate":      ["Salatbuffet"],
  "partypfanne": ["..."],
  "dessert":     ["Dessertbuffet"]
}

Regeln:
- Schreibe Speisen IMMER auf Deutsch und mit großem Anfangsbuchstaben
- Wenn eine Kategorie fehlt, nutze sinnvolle Standardwerte
- Salate → immer "Salatbuffet" (außer explizit anders angegeben)
- Dessert → immer "Dessertbuffet" (außer explizit anders angegeben)
- Partypfanne → wenn nicht genannt, schreibe "Tagesempfehlung"
- Pasta → EIGENE Kategorie für alle Nudelgerichte (Spaghetti, Tortellini,
  Lasagne, Penne, Tagliatelle, Cannelloni, Rigatoni, Spätzle mit Käse, ...).
  Nudelgerichte gehören NICHT zu "haupt", sondern in "pasta".
  Wenn keine Pasta genannt ist, gib eine leere Liste zurück: "pasta": []
- "haupt" enthält NUR Fleisch/Fisch/Geflügel-Hauptgerichte, KEINE Nudelgerichte
- Beilagen: Gemüse zuerst, dann Stärke
- Gib NUR das JSON zurück, sonst nichts"""


def ki_kategorisieren(text: str, api_key: str) -> dict:
    """Sendet Freitext an Claude und gibt das kategorisierte Menü zurück."""
    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=512,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": text}],
    )
    raw = message.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.strip()
    menu = json.loads(raw)

    defaults = {
        "suppe":       [],
        "haupt":       [],
        "pasta":       [],
        "beilagen":    [],
        "salate":      ["Salatbuffet"],
        "partypfanne": ["Tagesempfehlung"],
        "dessert":     ["Dessertbuffet"],
    }
    for key, default in defaults.items():
        if key not in menu:
            menu[key] = default
        elif not menu[key] and default:
            # nur überschreiben, wenn ein nicht-leerer Default existiert
            menu[key] = default
    return menu


# ══════════════════════════════════════════════════════════════════════════════
# Slot-Konfiguration der Comboboxen
# ══════════════════════════════════════════════════════════════════════════════
#
# Jeder Eintrag: (Combobox-Key, Kategorie-Key, Slot-Index, Label-Key, optional)
#
SLOTS = [
    ("suppe_1",      "suppe",       0, "cat_suppe",       False),
    ("haupt_1",      "haupt",       0, "cat_haupt",       False),
    ("haupt_2",      "haupt",       1, "cat_haupt",       False),
    ("pasta_1",      "pasta",       0, "cat_pasta",       True),
    ("beilage_1",    "beilagen",    0, "cat_beilage",     False),
    ("beilage_2",    "beilagen",    1, "cat_beilage",     False),
    ("beilage_3",    "beilagen",    2, "cat_beilage",     False),
    ("beilage_4",    "beilagen",    3, "cat_beilage",     True),
    ("salate_1",     "salate",      0, "cat_salate",      False),
    ("partypfanne_1","partypfanne", 0, "cat_partypfanne", False),
    ("dessert_1",    "dessert",     0, "cat_dessert",     False),
]

# Mapping Excel-Label (Großbuchstaben) → Kategorie-Key
LABEL_TO_KEY = {label: key for key, label, *_ in CATEGORIES}


# ══════════════════════════════════════════════════════════════════════════════
# Vorschläge aus Excel-Archiv (Top 10 pro Kategorie)
# ══════════════════════════════════════════════════════════════════════════════

DEFAULT_SUGGESTIONS = {
    "suppe": [
        "Tomatensuppe", "Hühnerbrühe mit Nudeln", "Kürbissuppe", "Linsensuppe",
        "Spargelcremesuppe", "Gulaschsuppe", "Kartoffelsuppe", "Erbsensuppe",
        "Brokkolicremesuppe", "Tagessuppe",
    ],
    "haupt": [
        "Wiener Schnitzel", "Kasseler", "Hähnchenbrust", "Rinderbraten",
        "Putenrollbraten", "Schweinebraten", "Frikadellen", "Lachsfilet",
        "Hähnchenschenkel", "Rouladen",
    ],
    "pasta": [
        "Spaghetti Bolognese", "Tortellini in Sahnesauce", "Penne Arrabbiata",
        "Lasagne", "Spaghetti Carbonara", "Tagliatelle mit Lachs",
        "Nudelauflauf", "Rigatoni mit Tomatensauce", "Cannelloni",
        "Spätzle mit Käse",
    ],
    "beilagen": [
        "Bratkartoffeln", "Salzkartoffeln", "Kartoffelpüree", "Reis",
        "Spätzle", "Pommes frites", "Brokkoli", "Erbsen und Möhren",
        "Blumenkohl", "Rotkohl",
    ],
    "salate": [
        "Salatbuffet", "Gemischter Salat", "Tomatensalat", "Gurkensalat",
        "Krautsalat", "Kartoffelsalat", "Nudelsalat", "Heringssalat",
        "Bohnensalat", "Rote-Bete-Salat",
    ],
    "partypfanne": [
        "Tagesempfehlung", "Hähnchen Süß-Sauer", "Chili con Carne",
        "Bauernpfanne", "Gyrospfanne", "Geschnetzeltes Züricher Art",
        "Currywurst-Pfanne", "Nudelpfanne", "Reispfanne", "Gemüsepfanne",
    ],
    "dessert": [
        "Dessertbuffet", "Vanillepudding", "Schokopudding", "Obstsalat",
        "Mousse au Chocolat", "Rote Grütze", "Tiramisu", "Apfelstrudel",
        "Eis mit heißen Kirschen", "Quarkspeise",
    ],
}


def build_suggestions_from_excel(archiv_file: Path, top_n: int | None = None) -> dict:
    """
    Liest das Excel-Archiv und gibt ALLE jemals verwendeten Gerichte pro Kategorie zurück.
    Sortierung: nach Häufigkeit absteigend, dann alphabetisch.
    DEFAULT_SUGGESTIONS werden als Sockel angehängt (falls noch nicht enthalten),
    damit auch bei leerem Archiv etwas zur Auswahl steht.
    """
    base_defaults = {k: list(v) for k, v in DEFAULT_SUGGESTIONS.items()}
    # Fehlende Kategorien (z. B. neu hinzugekommene wie "pasta") absichern
    for key, *_ in CATEGORIES:
        base_defaults.setdefault(key, [])

    if not (OPENPYXL_OK and archiv_file.exists()):
        result = base_defaults
        return {k: (v[:top_n] if top_n else v) for k, v in result.items()}

    try:
        wb = load_workbook(archiv_file, read_only=True, data_only=True)
        ws = wb.active
        counts: dict = {key: collections.Counter() for key, *_ in CATEGORIES}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            kategorie, speise = row[1], row[2]
            if not kategorie or not speise:
                continue
            key = LABEL_TO_KEY.get(str(kategorie).strip().upper())
            if key:
                counts[key][str(speise).strip()] += 1
        wb.close()

        result: dict = {}
        for key, *_ in CATEGORIES:
            # Nach Häufigkeit, dann alphabetisch
            items = sorted(counts[key].items(), key=lambda kv: (-kv[1], kv[0].lower()))
            names = [n for n, _ in items]
            # Defaults als Sockel anhängen (für neue/leere Kategorien)
            for d in base_defaults.get(key, []):
                if d not in names:
                    names.append(d)
            result[key] = names[:top_n] if top_n else names
        return result
    except Exception as e:
        print(f"Warnung: Vorschläge aus Excel konnten nicht gelesen werden – {e}")
        return {k: (v[:top_n] if top_n else v) for k, v in base_defaults.items()}


# ══════════════════════════════════════════════════════════════════════════════
# Kalender-Widget (Fallback)
# ══════════════════════════════════════════════════════════════════════════════

class SimpleCalendar(tk.Toplevel):
    WEEKDAYS = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
    MONTHS_DE = [
        "", "Januar", "Februar", "März", "April", "Mai", "Juni",
        "Juli", "August", "September", "Oktober", "November", "Dezember",
    ]

    def __init__(self, parent, callback, initial_date: datetime):
        super().__init__(parent)
        self.callback     = callback
        self.current_date = initial_date
        self.view_date    = datetime(initial_date.year, initial_date.month, 1)

        self.title("Datum wählen")
        self.resizable(False, False)
        self.configure(bg=C_WHITE)
        self.grab_set()
        self._build()
        self._render()

    def _build(self):
        nav = tk.Frame(self, bg=C_TEAL)
        nav.pack(fill="x")
        tk.Button(nav, text="◀", font=FONT_LABEL, bg=C_TEAL, fg=C_WHITE,
                  bd=0, command=self._prev_month).pack(side="left", padx=8, pady=6)
        self.lbl_month = tk.Label(nav, text="", font=FONT_LABEL,
                                  bg=C_TEAL, fg=C_WHITE, width=18)
        self.lbl_month.pack(side="left", expand=True)
        tk.Button(nav, text="▶", font=FONT_LABEL, bg=C_TEAL, fg=C_WHITE,
                  bd=0, command=self._next_month).pack(side="right", padx=8, pady=6)
        hdr = tk.Frame(self, bg=C_WHITE)
        hdr.pack(fill="x", padx=8, pady=(8, 0))
        for i, d in enumerate(self.WEEKDAYS):
            color = C_TEAL_DARK if i >= 5 else C_GRAY
            tk.Label(hdr, text=d, font=FONT_SMALL, fg=color, bg=C_WHITE,
                     width=4).grid(row=0, column=i)
        self.grid_frame = tk.Frame(self, bg=C_WHITE)
        self.grid_frame.pack(padx=8, pady=4)

    def _render(self):
        self.lbl_month.config(text=f"{self.MONTHS_DE[self.view_date.month]} {self.view_date.year}")
        for w in self.grid_frame.winfo_children():
            w.destroy()
        first_weekday = self.view_date.weekday()
        if self.view_date.month == 12:
            days = (datetime(self.view_date.year + 1, 1, 1) - self.view_date).days
        else:
            days = (datetime(self.view_date.year, self.view_date.month + 1, 1) - self.view_date).days
        col, row = first_weekday, 0
        for day in range(1, days + 1):
            d = datetime(self.view_date.year, self.view_date.month, day)
            is_selected = (d.date() == self.current_date.date())
            is_weekend  = d.weekday() >= 5
            bg = C_TEAL if is_selected else C_WHITE
            fg = C_WHITE if is_selected else (C_TEAL_DARK if is_weekend else C_TEXT)
            tk.Button(self.grid_frame, text=str(day), width=3,
                      font=FONT_SMALL, bg=bg, fg=fg, bd=0, relief="flat",
                      cursor="hand2",
                      command=lambda dt=d: self._select(dt)
                      ).grid(row=row, column=col, padx=1, pady=1)
            col += 1
            if col > 6:
                col = 0
                row += 1

    def _prev_month(self):
        if self.view_date.month == 1:
            self.view_date = datetime(self.view_date.year - 1, 12, 1)
        else:
            self.view_date = datetime(self.view_date.year, self.view_date.month - 1, 1)
        self._render()

    def _next_month(self):
        if self.view_date.month == 12:
            self.view_date = datetime(self.view_date.year + 1, 1, 1)
        else:
            self.view_date = datetime(self.view_date.year, self.view_date.month + 1, 1)
        self._render()

    def _select(self, date: datetime):
        self.current_date = date
        self.callback(date)
        self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
# Vollansicht-Fenster
# ══════════════════════════════════════════════════════════════════════════════

class VorschauFenster(tk.Toplevel):
    """Großes Vorschaufenster (öffnet das gespeicherte Bild)."""

    def __init__(self, parent, image_path: Path, title_text: str, btn_text: str):
        super().__init__(parent)
        self.title(f"{title_text} – {image_path.name}")
        self.configure(bg=C_BG)
        self.image_path = image_path

        screen_h = self.winfo_screenheight()
        target_h = int(screen_h * 0.80)
        target_w = int(target_h * 1080 / 1920)
        self.geometry(f"{target_w + 20}x{target_h + 60}")
        self.resizable(True, True)

        bar = tk.Frame(self, bg=C_TEAL, height=40)
        bar.pack(fill="x")
        bar.pack_propagate(False)
        tk.Label(bar, text=f"📷  {title_text}", font=FONT_LABEL,
                 bg=C_TEAL, fg=C_WHITE).pack(side="left", padx=12, pady=8)
        tk.Button(bar, text=btn_text, font=FONT_SMALL,
                  bg=C_TEAL_DARK, fg=C_WHITE, bd=0, padx=8,
                  command=self._open_folder).pack(side="right", padx=8, pady=6)

        if PIL_OK:
            try:
                img = Image.open(self.image_path)
                img = img.resize((target_w, target_h), Image.LANCZOS)
                self._photo = ImageTk.PhotoImage(img)
                canvas = tk.Canvas(self, width=target_w, height=target_h,
                                   bg=C_BG, highlightthickness=0)
                canvas.pack(padx=10, pady=10)
                canvas.create_image(0, 0, anchor="nw", image=self._photo)
            except Exception as e:
                tk.Label(self, text=f"Bild konnte nicht geladen werden:\n{e}",
                         font=FONT_NORMAL, bg=C_BG, fg=C_ERROR).pack(pady=40)

    def _open_folder(self):
        folder = str(self.image_path.parent)
        if sys.platform == "win32":
            os.startfile(folder)
        elif sys.platform == "darwin":
            os.system(f"open '{folder}'")
        else:
            os.system(f"xdg-open '{folder}'")


# ══════════════════════════════════════════════════════════════════════════════
# Hauptfenster
# ══════════════════════════════════════════════════════════════════════════════

class TagesbuffetGUI:
    """Haupt-GUI-Klasse."""

    WEEKDAYS_DE = ["Montag", "Dienstag", "Mittwoch", "Donnerstag",
                   "Freitag", "Samstag", "Sonntag"]
    MONTHS_DE   = ["", "Januar", "Februar", "März", "April", "Mai", "Juni",
                   "Juli", "August", "September", "Oktober", "November", "Dezember"]

    DEBOUNCE_MS = 300   # Wartezeit nach letzter Änderung bis Re-Render

    def __init__(self):
        self.root = tk.Tk()
        self.root.configure(bg=C_BG)
        self.root.minsize(1180, 880)

        # ── Pfade & Settings ─────────────────────────────────────────────────
        default_dir = Path(__file__).parent / "buffet"
        default_dir.mkdir(parents=True, exist_ok=True)

        self.image_dir:  Path = default_dir
        self.archiv_dir: Path = default_dir
        self.lang: str        = "de"

        self._load_settings()

        # ── State ────────────────────────────────────────────────────────────
        self.selected_date              = datetime.today()
        self.last_img_path: Optional[Path] = None
        self._render_after: Optional[str]  = None
        self._preview_photo                = None  # PIL ImageTk-Ref behalten
        self._suggestions:  dict           = {}
        self._translatable: list           = []   # [(widget, key, attr)]
        self._slot_label_widgets: dict     = {}   # slot_key → Label-Widget

        self.api_key = tk.StringVar(value=os.getenv("ANTHROPIC_API_KEY", ""))

        # Generator
        self.generator = (
            TagesbuffetGenerator(image_dir=self.image_dir, archiv_dir=self.archiv_dir)
            if GENERATOR_OK else None
        )

        # Vorschläge laden
        self._suggestions = build_suggestions_from_excel(self.archiv_dir / "Tagesbuffet_Archiv.xlsx")

        # ── UI aufbauen ──────────────────────────────────────────────────────
        self._build_ui()
        self._apply_language()
        self._load_history()
        self._set_combobox_values()

        # Standardwerte für Salate/Dessert vorbelegen
        self._preset_defaults()

        # Erste Vorschau rendern
        self._schedule_render()

        # Warnung, falls gespeicherte Pfade nicht (mehr) existieren
        if getattr(self, "_missing_paths", None):
            names = ", ".join(label for label, _ in self._missing_paths)
            self._set_status(
                f"⚠ Gespeicherter Ordner ({names}) nicht gefunden – Default verwendet.",
                C_ERROR,
            )

        # Icon
        icon_path = Path(__file__).parent / "buffet_icon.ico"
        if icon_path.exists():
            try:
                self.root.iconbitmap(str(icon_path))
            except Exception:
                pass

    # ── Sprache / i18n ────────────────────────────────────────────────────────

    def t(self, key: str, **fmt) -> str:
        """Übersetzt Key in aktuelle Sprache. Optional mit .format(**fmt)."""
        text = TRANSLATIONS.get(self.lang, TRANSLATIONS["de"]).get(key, key)
        if fmt:
            try:
                return text.format(**fmt)
            except Exception:
                return text
        return text

    def _register(self, widget, key: str, attr: str = "text", **fmt):
        """Merkt sich, dass widget[attr] bei Sprachwechsel neu gesetzt werden soll."""
        self._translatable.append((widget, key, attr, fmt))
        try:
            widget[attr] = self.t(key, **fmt)
        except Exception:
            widget.config(**{attr: self.t(key, **fmt)})

    def _apply_language(self):
        """Setzt alle registrierten Texte neu (nach Sprachwechsel)."""
        self.root.title(self.t("app_title"))
        for widget, key, attr, fmt in self._translatable:
            try:
                widget[attr] = self.t(key, **fmt)
            except (tk.TclError, KeyError):
                pass
        # Datums-Label aktualisieren
        if hasattr(self, "lbl_datum"):
            self.lbl_datum.config(text=self._format_date(self.selected_date))
        # Pfad-Texte aktualisieren
        self._update_path_labels()
        # Verlauf neu rendern (wegen "klicken zum übernehmen" etc.)
        self._load_history()

    def _toggle_language(self):
        self.lang = "ar" if self.lang == "de" else "de"
        self._save_settings()
        self._apply_language()

    # ── UI-Aufbau ─────────────────────────────────────────────────────────────

    def _build_ui(self):
        self._build_header()

        # Hauptbereich
        main = tk.Frame(self.root, bg=C_BG)
        main.pack(fill="both", expand=True, padx=12, pady=8)

        left  = tk.Frame(main, bg=C_BG, width=720)
        left.pack(side="left", fill="both", padx=(0, 8))
        left.pack_propagate(False)

        right = tk.Frame(main, bg=C_BG)
        right.pack(side="left", fill="both", expand=True)

        # LINKS
        self._build_datum_row(left)
        self._build_freitext(left)
        self._build_menu_grid(left)
        self._build_apikey(left)
        self._build_status(left)

        # RECHTS
        self._build_preview_panel(right)

    def _build_header(self):
        header = tk.Frame(self.root, bg=C_TEAL, height=80)
        header.pack(fill="x")
        header.pack_propagate(False)

        # Logo links
        if PIL_OK and Path(LOGO_PATH).exists():
            try:
                logo_raw = Image.open(LOGO_PATH).convert("RGBA")
                ratio    = 56 / logo_raw.height
                logo_w   = int(logo_raw.width * ratio)
                logo_raw = logo_raw.resize((logo_w, 56), Image.LANCZOS)
                bg_img = Image.new("RGBA", (logo_w, 56), (*self._hex_to_rgb(C_TEAL), 255))
                bg_img.paste(logo_raw, mask=logo_raw.split()[3])
                self._header_logo = ImageTk.PhotoImage(bg_img.convert("RGB"))
                tk.Label(header, image=self._header_logo, bg=C_TEAL
                         ).pack(side="left", padx=(12, 6), pady=12)
            except Exception:
                pass

        # Titel + Untertitel
        ttl_frame = tk.Frame(header, bg=C_TEAL)
        ttl_frame.pack(side="left", padx=(4, 16), pady=12)

        self.lbl_header_title = tk.Label(
            ttl_frame, font=("Segoe UI", 16, "bold"),
            bg=C_TEAL, fg=C_WHITE, anchor="w"
        )
        self.lbl_header_title.pack(anchor="w")
        self._register(self.lbl_header_title, "header_title")

        self.lbl_header_sub = tk.Label(
            ttl_frame, font=FONT_SMALL,
            bg=C_TEAL, fg="#d0f0f0", anchor="w"
        )
        self.lbl_header_sub.pack(anchor="w")
        self._register(self.lbl_header_sub, "header_subtitle")

        # Sprachschalter ganz rechts
        self.btn_lang = tk.Button(
            header, font=FONT_SMALL,
            bg=C_TEAL_DARK, fg=C_WHITE, bd=0, padx=10, pady=4,
            cursor="hand2", command=self._toggle_language,
        )
        self.btn_lang.pack(side="right", padx=(6, 12), pady=22)
        self._register(self.btn_lang, "lang_switch")

        # Pfad-Buttons (Excel + Bilder) NEBENEINANDER
        path_frame = tk.Frame(header, bg=C_TEAL)
        path_frame.pack(side="right", padx=(6, 0), pady=18)

        # Excel (links)
        self.btn_path_excel = tk.Button(
            path_frame, font=FONT_SMALL,
            bg=C_WHITE, fg=C_TEAL_DARK, bd=0, padx=8, pady=2,
            cursor="hand2", command=self._change_archiv_dir,
        )
        self.btn_path_excel.pack(side="left")
        self._register(self.btn_path_excel, "btn_path_excel")
        self.lbl_path_excel = tk.Label(
            path_frame, font=FONT_TINY, bg=C_TEAL, fg=C_WHITE,
            cursor="hand2", anchor="w",
        )
        self.lbl_path_excel.pack(side="left", padx=(6, 12))
        self.lbl_path_excel.bind("<Button-1>", lambda e: self._open_folder(self.archiv_dir))

        # Bilder (rechts)
        self.btn_path_images = tk.Button(
            path_frame, font=FONT_SMALL,
            bg=C_WHITE, fg=C_TEAL_DARK, bd=0, padx=8, pady=2,
            cursor="hand2", command=self._change_image_dir,
        )
        self.btn_path_images.pack(side="left")
        self._register(self.btn_path_images, "btn_path_images")
        self.lbl_path_images = tk.Label(
            path_frame, font=FONT_TINY, bg=C_TEAL, fg=C_WHITE,
            cursor="hand2", anchor="w",
        )
        self.lbl_path_images.pack(side="left", padx=(6, 0))
        self.lbl_path_images.bind("<Button-1>", lambda e: self._open_folder(self.image_dir))

        self._update_path_labels()

    def _build_datum_row(self, parent):
        frame = tk.LabelFrame(parent, font=FONT_IN_FRAME,
                               bg=C_BG, fg=C_TEAL_DARK, bd=1, relief="groove")
        frame.pack(fill="x", pady=(0, 6))
        self._register(frame, "frame_date")

        inner = tk.Frame(frame, bg=C_BG)
        inner.pack(fill="x", padx=8, pady=6)

        self.lbl_datum = tk.Label(
            inner, text=self._format_date(self.selected_date),
            font=FONT_IN_DATE, bg=C_BG, fg=C_TEAL_DARK,
        )
        self.lbl_datum.pack(side="left")

        btn_frame = tk.Frame(inner, bg=C_BG)
        btn_frame.pack(side="right")

        b1 = tk.Button(btn_frame, font=FONT_IN_BTN,
                       bg=C_TEAL_LIGHT, fg=C_TEAL_DARK, bd=1, relief="solid",
                       command=lambda: self._set_date(datetime.today()))
        b1.pack(side="left", padx=2)
        self._register(b1, "btn_today")

        b2 = tk.Button(btn_frame, font=FONT_IN_BTN,
                       bg=C_TEAL_LIGHT, fg=C_TEAL_DARK, bd=1, relief="solid",
                       command=lambda: self._set_date(datetime.today() + timedelta(days=1)))
        b2.pack(side="left", padx=2)
        self._register(b2, "btn_tomorrow")

        b3 = tk.Button(btn_frame, font=FONT_IN_BTN,
                       bg=C_TEAL, fg=C_WHITE, bd=0,
                       command=self._open_calendar)
        b3.pack(side="left", padx=(6, 0))
        self._register(b3, "btn_calendar")

    def _build_freitext(self, parent):
        frame = tk.LabelFrame(parent, font=FONT_IN_FRAME,
                               bg=C_BG, fg=C_TEAL_DARK, bd=1, relief="groove")
        frame.pack(fill="x", pady=(0, 6))
        self._register(frame, "frame_freitext")

        self.lbl_hint_freitext = tk.Label(
            frame, font=FONT_IN_SMALL, bg=C_BG, fg=C_GRAY,
        )
        self.lbl_hint_freitext.pack(anchor="w", padx=8, pady=(4, 2))
        self._register(self.lbl_hint_freitext, "hint_freitext")

        self.txt_eingabe = scrolledtext.ScrolledText(
            frame, height=3, font=FONT_IN_NORMAL,
            bg=C_WHITE, fg=C_TEXT, wrap="word", bd=1, relief="solid",
            insertbackground=C_TEAL_DARK,
        )
        self.txt_eingabe.pack(fill="x", padx=8, pady=(0, 4))
        self._set_placeholder()

        self.txt_eingabe.bind("<FocusIn>",  self._freitext_focus_in)
        self.txt_eingabe.bind("<FocusOut>", self._freitext_focus_out)

        self.btn_analyse = tk.Button(
            frame, font=FONT_IN_KI,
            bg=C_TEAL, fg=C_WHITE, bd=0, padx=10, pady=8, cursor="hand2",
            command=self._start_analyse,
        )
        self.btn_analyse.pack(fill="x", padx=8, pady=(0, 8))
        self._register(self.btn_analyse, "btn_ki")

    def _build_menu_grid(self, parent):
        frame = tk.LabelFrame(parent, font=FONT_IN_FRAME,
                               bg=C_BG, fg=C_TEAL_DARK, bd=1, relief="groove")
        frame.pack(fill="both", expand=True, pady=(0, 6))
        self._register(frame, "frame_menu")

        inner = tk.Frame(frame, bg=C_BG)
        inner.pack(fill="x", padx=8, pady=6)

        # Combobox-Dropdown-Liste auf große Schrift einstellen
        # (das Eingabefeld der Combobox bekommt font= direkt am Widget)
        self.root.option_add("*TCombobox*Listbox.font", FONT_IN_NORMAL)

        self.combo_vars: dict = {}
        self.comboboxes:  dict = {}

        # Numbering pro Kategorie für Label "Hauptgericht 1/2/3"
        cat_counters: dict = {}

        for slot_key, cat_key, slot_idx, label_key, optional in SLOTS:
            row = tk.Frame(inner, bg=C_BG)
            row.pack(fill="x", pady=4)

            n = cat_counters.get(cat_key, 0) + 1
            cat_counters[cat_key] = n

            lbl = tk.Label(row, font=FONT_IN_LABEL, fg=C_TEAL_DARK,
                           bg=C_BG, width=20, anchor="w")
            lbl.pack(side="left")
            # Label-Text mit Nummerierung (für Kategorien mit mehreren Slots)
            if "{n}" in self.t(label_key):
                self._register(lbl, label_key, n=n)
            else:
                self._register(lbl, label_key)
            self._slot_label_widgets[slot_key] = (lbl, label_key, n)

            var = tk.StringVar()
            # height = sichtbare Zeilen im Dropdown; bei mehr Einträgen erscheint
            # automatisch eine Scrollbar.
            cb = ttk.Combobox(row, textvariable=var, font=FONT_IN_NORMAL,
                              state="normal", width=28, height=20)
            cb.pack(side="left", fill="x", expand=True)
            cb.bind("<<ComboboxSelected>>", lambda e: self._schedule_render())
            cb.bind("<KeyRelease>",        lambda e: self._schedule_render())
            cb.bind("<FocusOut>",          lambda e: self._schedule_render())

            # Optional-Tag
            if optional:
                opt = tk.Label(row, font=FONT_IN_SMALL, fg=C_GRAY, bg=C_BG)
                opt.pack(side="left", padx=(4, 0))
                self._register(opt, "hint_optional")

            self.combo_vars[slot_key] = var
            self.comboboxes[slot_key]  = cb

    def _build_apikey(self, parent):
        frame = tk.LabelFrame(parent, font=FONT_IN_FRAME,
                               bg=C_BG, fg=C_TEAL_DARK, bd=1, relief="groove")
        frame.pack(fill="x", pady=(0, 6))
        self._register(frame, "frame_apikey")

        inner = tk.Frame(frame, bg=C_BG)
        inner.pack(fill="x", padx=8, pady=6)

        entry = tk.Entry(inner, textvariable=self.api_key,
                         font=FONT_IN_SMALL, show="•",
                         bg=C_WHITE, fg=C_TEXT, bd=1, relief="solid")
        entry.pack(side="left", fill="x", expand=True)

        self.lbl_apikey_hint = tk.Label(inner, font=FONT_IN_SMALL,
                                         bg=C_BG, fg=C_GRAY)
        self.lbl_apikey_hint.pack(side="left", padx=(8, 0))
        self._register(self.lbl_apikey_hint, "lbl_apikey_hint")

    def _build_status(self, parent):
        self.lbl_status = tk.Label(
            parent, font=FONT_IN_SMALL, bg=C_BG, fg=C_GRAY, anchor="w",
        )
        self.lbl_status.pack(fill="x", pady=(2, 0))
        self._register(self.lbl_status, "status_ready")

    def _build_preview_panel(self, parent):
        # Vorschau
        prev_frame = tk.LabelFrame(parent, font=FONT_LABEL,
                                    bg=C_BG, fg=C_TEAL_DARK, bd=1, relief="groove")
        prev_frame.pack(fill="both", expand=True)
        self._register(prev_frame, "frame_preview")

        # Preview-Größe: schmaler als zuvor (Verhältnis 1080:1920)
        self.PREVIEW_W = 320
        self.PREVIEW_H = 569

        self.preview_canvas = tk.Canvas(
            prev_frame, width=self.PREVIEW_W, height=self.PREVIEW_H,
            bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER,
        )
        self.preview_canvas.pack(padx=10, pady=10)

        # Aktions-Buttons
        btn_frame = tk.Frame(parent, bg=C_BG)
        btn_frame.pack(fill="x", pady=(8, 8))

        self.btn_save = tk.Button(
            btn_frame, font=FONT_LABEL,
            bg=C_TEAL_DARK, fg=C_WHITE, bd=0, padx=10, pady=8, cursor="hand2",
            command=self._save_image,
        )
        self.btn_save.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self._register(self.btn_save, "btn_save_image")

        self.btn_excel = tk.Button(
            btn_frame, font=FONT_LABEL,
            bg=C_SUCCESS, fg=C_WHITE, bd=0, padx=10, pady=8, cursor="hand2",
            command=self._save_excel,
        )
        self.btn_excel.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self._register(self.btn_excel, "btn_save_excel")

        self.btn_full = tk.Button(
            btn_frame, font=FONT_LABEL,
            bg="#5c6bc0", fg=C_WHITE, bd=0, padx=10, pady=8, cursor="hand2",
            command=self._show_full_preview,
        )
        self.btn_full.pack(side="left", fill="x", expand=True)
        self._register(self.btn_full, "btn_full_preview")

        # Verlauf (kompakt)
        hist_frame = tk.LabelFrame(parent, font=FONT_LABEL,
                                    bg=C_BG, fg=C_TEAL_DARK, bd=1, relief="groove")
        hist_frame.pack(fill="x")
        self._register(hist_frame, "frame_history")

        self.lbl_history_hint = tk.Label(
            hist_frame, font=FONT_TINY, bg=C_BG, fg=C_GRAY,
        )
        self.lbl_history_hint.pack(anchor="w", padx=8, pady=(4, 2))
        self._register(self.lbl_history_hint, "history_hint")

        self.history_container = tk.Frame(hist_frame, bg=C_BG, height=140)
        self.history_container.pack(fill="x", padx=4, pady=4)

    # ── Settings (Pfade + Sprache) ────────────────────────────────────────────

    def _settings_file(self) -> Path:
        return self.archiv_dir / ".gui_settings.json"

    def _load_settings(self):
        # Settings können in ./buffet liegen ODER (wenn einmal gewechselt) wo anders.
        # Wir suchen zuerst im Default-Ordner nach einer Settings-Datei mit Pfad-Verweis.
        self._missing_paths: list = []  # für Warnung in Statuszeile
        local_settings = Path(__file__).parent / "buffet" / ".gui_settings.json"
        try:
            if local_settings.exists():
                with open(local_settings, encoding="utf-8") as f:
                    data = json.load(f)
                if "image_dir" in data:
                    p = Path(data["image_dir"])
                    if p.exists():
                        self.image_dir = p
                    else:
                        self._missing_paths.append(("Bilder", str(p)))
                if "archiv_dir" in data:
                    p = Path(data["archiv_dir"])
                    if p.exists():
                        self.archiv_dir = p
                    else:
                        self._missing_paths.append(("Excel", str(p)))
                if "lang" in data and data["lang"] in TRANSLATIONS:
                    self.lang = data["lang"]
        except Exception as e:
            print(f"Warnung: Settings konnten nicht geladen werden – {e}")

    def _save_settings(self):
        # IMMER auch im lokalen Default-Ordner speichern, damit beim nächsten Start
        # die Pfade gefunden werden (auch wenn archiv_dir extern liegt).
        data = {
            "image_dir":  str(self.image_dir),
            "archiv_dir": str(self.archiv_dir),
            "lang":       self.lang,
        }
        for path in {Path(__file__).parent / "buffet", self.archiv_dir}:
            try:
                path.mkdir(parents=True, exist_ok=True)
                with open(path / ".gui_settings.json", "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            except Exception:
                pass

    # ── Pfad-Wechsel ──────────────────────────────────────────────────────────

    def _change_image_dir(self):
        p = filedialog.askdirectory(
            title=self.t("msg_dialog_image_title"),
            initialdir=str(self.image_dir),
        )
        if p:
            self.image_dir = Path(p)
            if self.generator:
                self.generator.set_image_dir(self.image_dir)
            self._save_settings()
            self._update_path_labels()
            self._set_status(self.t("status_path_changed"), C_SUCCESS)

    def _change_archiv_dir(self):
        p = filedialog.askdirectory(
            title=self.t("msg_dialog_excel_title"),
            initialdir=str(self.archiv_dir),
        )
        if p:
            self.archiv_dir = Path(p)
            if self.generator:
                self.generator.set_archiv_dir(self.archiv_dir)
            self._save_settings()
            self._update_path_labels()
            # Vorschläge + Verlauf neu laden
            self._suggestions = build_suggestions_from_excel(
                self.archiv_dir / "Tagesbuffet_Archiv.xlsx"
            )
            self._set_combobox_values()
            self._load_history()
            self._set_status(self.t("status_path_changed"), C_SUCCESS)

    def _update_path_labels(self):
        if hasattr(self, "lbl_path_images"):
            self.lbl_path_images.config(text=self._shorten_path(self.image_dir))
        if hasattr(self, "lbl_path_excel"):
            self.lbl_path_excel.config(text=self._shorten_path(self.archiv_dir))

    @staticmethod
    def _shorten_path(p: Path, max_len: int = 38) -> str:
        s = str(p)
        if len(s) <= max_len:
            return s
        # Zeige Anfang und Ende
        return s[:14] + " … " + s[-(max_len - 17):]

    @staticmethod
    def _open_folder(path: Path):
        try:
            path.mkdir(parents=True, exist_ok=True)
            if sys.platform == "win32":
                os.startfile(str(path))
            elif sys.platform == "darwin":
                os.system(f"open '{path}'")
            else:
                os.system(f"xdg-open '{path}'")
        except Exception as e:
            print(f"Konnte Ordner nicht öffnen: {e}")

    # ── Combobox-Werte ────────────────────────────────────────────────────────

    def _set_combobox_values(self):
        for slot_key, cat_key, *_ in SLOTS:
            cb = self.comboboxes.get(slot_key)
            if cb is not None:
                cb["values"] = self._suggestions.get(cat_key, [])

    def _preset_defaults(self):
        # Standardwerte aus CATEGORIES (z. B. Salatbuffet, Dessertbuffet)
        for key, label, _, _, default in CATEGORIES:
            if not default:
                continue
            for slot_key, cat_key, slot_idx, *_ in SLOTS:
                if cat_key == key and slot_idx < len(default):
                    if not self.combo_vars[slot_key].get():
                        self.combo_vars[slot_key].set(default[slot_idx])

    # ── Datum ─────────────────────────────────────────────────────────────────

    def _format_date(self, date: datetime) -> str:
        return f"{self.WEEKDAYS_DE[date.weekday()]}, {date.day:02d}. {self.MONTHS_DE[date.month]} {date.year}"

    def _set_date(self, date: datetime):
        self.selected_date = date
        self.lbl_datum.config(text=self._format_date(date))
        self._schedule_render()

    def _open_calendar(self):
        SimpleCalendar(self.root, self._set_date, self.selected_date)

    # ── Status ────────────────────────────────────────────────────────────────

    def _set_status(self, msg: str, color: str = C_GRAY):
        # Status-Texte werden direkt gesetzt (sind nicht mehr i18n-Keys)
        self.lbl_status.config(text=msg, fg=color)
        # Aus _translatable entfernen, damit Sprachwechsel ihn nicht überschreibt
        self._translatable = [
            (w, k, a, f) for (w, k, a, f) in self._translatable
            if w is not self.lbl_status
        ]
        self.root.update_idletasks()

    # ── Freitext-Placeholder ──────────────────────────────────────────────────

    def _set_placeholder(self):
        self.txt_eingabe.delete("1.0", "end")
        self.txt_eingabe.insert("1.0", self.t("placeholder_freitext"))
        self.txt_eingabe.config(fg=C_GRAY)
        self._freitext_is_placeholder = True

    def _freitext_focus_in(self, _e=None):
        if getattr(self, "_freitext_is_placeholder", False):
            self.txt_eingabe.delete("1.0", "end")
            self.txt_eingabe.config(fg=C_TEXT)
            self._freitext_is_placeholder = False

    def _freitext_focus_out(self, _e=None):
        if not self.txt_eingabe.get("1.0", "end-1c").strip():
            self._set_placeholder()

    def _get_eingabe_text(self) -> str:
        if getattr(self, "_freitext_is_placeholder", False):
            return ""
        return self.txt_eingabe.get("1.0", "end-1c").strip()

    # ── Menü aus den Comboboxen lesen / schreiben ────────────────────────────

    def _menu_aus_feldern(self) -> dict:
        menu: dict = {key: [] for key, *_ in CATEGORIES}
        for slot_key, cat_key, slot_idx, *_ in SLOTS:
            val = self.combo_vars[slot_key].get().strip()
            if val:
                menu[cat_key].append(val)
        return menu

    def _menu_in_felder(self, menu: dict):
        # Slots leeren, dann auffüllen
        for slot_key in self.combo_vars:
            self.combo_vars[slot_key].set("")
        for slot_key, cat_key, slot_idx, *_ in SLOTS:
            items = menu.get(cat_key, [])
            if slot_idx < len(items):
                self.combo_vars[slot_key].set(items[slot_idx])

    # ── KI-Analyse ────────────────────────────────────────────────────────────

    def _start_analyse(self):
        text = self._get_eingabe_text()
        if not text:
            messagebox.showwarning(self.t("msg_no_input_title"), self.t("msg_no_input"))
            return
        if not ANTHROPIC_OK:
            messagebox.showerror(self.t("msg_no_anthropic_title"), self.t("msg_no_anthropic"))
            return
        key = self.api_key.get().strip()
        if not key:
            messagebox.showwarning(self.t("msg_no_apikey_title"), self.t("msg_no_apikey"))
            return

        self.btn_analyse.config(state="disabled", text=self.t("btn_ki_running"))
        self._set_status(self.t("status_ki_running"), C_TEAL)

        def run():
            try:
                menu = ki_kategorisieren(text, key)
                self.root.after(0, lambda: self._analyse_done(menu))
            except Exception as e:
                err = str(e)
                self.root.after(0, lambda: self._analyse_error(err))

        threading.Thread(target=run, daemon=True).start()

    def _analyse_done(self, menu: dict):
        self._menu_in_felder(menu)
        self.btn_analyse.config(state="normal", text=self.t("btn_ki"))
        self._set_status(self.t("status_ki_ok"), C_SUCCESS)
        self._schedule_render()

    def _analyse_error(self, err: str):
        self.btn_analyse.config(state="normal", text=self.t("btn_ki"))
        self._set_status(self.t("msg_ki_error", err=err), C_ERROR)
        messagebox.showerror(self.t("msg_ki_error_title"),
                             self.t("msg_ki_error", err=err))

    # ── Live-Vorschau ─────────────────────────────────────────────────────────

    def _schedule_render(self):
        """Plant ein Re-Render mit Debounce."""
        if self._render_after is not None:
            try:
                self.root.after_cancel(self._render_after)
            except Exception:
                pass
        self._render_after = self.root.after(self.DEBOUNCE_MS, self._render_preview)

    def _render_preview(self):
        self._render_after = None
        if not self.generator:
            return
        menu = self._menu_aus_feldern()
        try:
            img = self.generator.render_image(self.selected_date, menu)
            # Skalieren auf Vorschaugröße
            img_small = img.resize((self.PREVIEW_W, self.PREVIEW_H), Image.LANCZOS)
            self._preview_photo = ImageTk.PhotoImage(img_small)
            self.preview_canvas.delete("all")
            self.preview_canvas.create_image(0, 0, anchor="nw", image=self._preview_photo)
        except Exception as e:
            self.preview_canvas.delete("all")
            self.preview_canvas.create_text(
                self.PREVIEW_W // 2, self.PREVIEW_H // 2,
                text=f"Vorschau-Fehler:\n{e}", fill=C_ERROR, font=FONT_NORMAL,
            )

    # ── Speichern ─────────────────────────────────────────────────────────────

    def _save_image(self):
        if not self.generator:
            return
        menu = self._menu_aus_feldern()
        if not any(menu.values()):
            messagebox.showwarning(self.t("msg_menu_empty_title"),
                                   self.t("msg_menu_empty"))
            return
        try:
            path = self.generator.create_image(self.selected_date, menu)
            self.last_img_path = path
            # Excel automatisch mitschreiben (wie früher)
            try:
                self.generator.update_excel(self.selected_date, menu)
            except Exception as e:
                self._set_status(f"✓ Bild gespeichert – Excel-Fehler: {e}", C_ERROR)
                messagebox.showerror(self.t("msg_excel_error_title"), str(e))
                return
            self._save_to_history(self.selected_date, menu)
            # Vorschläge + Verlauf neu (häufig genutzte Gerichte ändern sich)
            self._suggestions = build_suggestions_from_excel(
                self.archiv_dir / "Tagesbuffet_Archiv.xlsx"
            )
            self._set_combobox_values()
            self._load_history()
            self._set_status(self.t("status_image_saved", name=path.name), C_SUCCESS)
        except Exception as e:
            self._set_status(f"Fehler: {e}", C_ERROR)
            messagebox.showerror(self.t("msg_image_error_title"),
                                 self.t("msg_image_error", err=str(e)))

    def _save_excel(self):
        if not self.generator:
            return
        menu = self._menu_aus_feldern()
        if not any(menu.values()):
            messagebox.showwarning(self.t("msg_menu_empty_title"),
                                   self.t("msg_menu_empty"))
            return
        try:
            self.generator.update_excel(self.selected_date, menu)
            self._set_status(self.t("status_excel_saved"), C_SUCCESS)
        except Exception as e:
            self._set_status(f"Fehler: {e}", C_ERROR)
            messagebox.showerror(self.t("msg_excel_error_title"), str(e))

    def _show_full_preview(self):
        if not self.last_img_path or not self.last_img_path.exists():
            messagebox.showinfo(self.t("msg_no_image_title"), self.t("msg_no_image"))
            return
        VorschauFenster(
            self.root, self.last_img_path,
            title_text=self.t("frame_preview").strip(),
            btn_text=self.t("msg_open_in_explorer"),
        )

    # ── Verlauf ───────────────────────────────────────────────────────────────

    def _history_file(self) -> Path:
        return self.archiv_dir / ".buffet_history.json"

    def _save_to_history(self, date: datetime, menu: dict):
        history = self._read_history()
        entry = {
            "date":    date.strftime("%d.%m.%Y"),
            "weekday": self.WEEKDAYS_DE[date.weekday()],
            "menu":    menu,
        }
        history = [h for h in history if h["date"] != entry["date"]]
        history.insert(0, entry)
        history = history[:14]
        try:
            self.archiv_dir.mkdir(parents=True, exist_ok=True)
            with open(self._history_file(), "w", encoding="utf-8") as f:
                json.dump(history, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _read_history(self) -> list:
        try:
            if self._history_file().exists():
                with open(self._history_file(), encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
        return []

    def _load_history(self):
        if not hasattr(self, "history_container"):
            return
        for w in self.history_container.winfo_children():
            w.destroy()

        history = self._read_history()[:7]
        if not history:
            tk.Label(self.history_container, text=self.t("history_empty"),
                     font=FONT_SMALL, bg=C_BG, fg=C_GRAY).pack(pady=8)
            return

        # Horizontal: max. 7 Karten in einer Reihe
        for i, entry in enumerate(history):
            self._build_history_card(entry, column=i)

    def _build_history_card(self, entry: dict, column: int):
        card = tk.Frame(self.history_container, bg=C_WHITE, bd=1, relief="solid",
                        cursor="hand2", width=110)
        card.grid(row=0, column=column, padx=2, pady=2, sticky="ns")
        card.grid_propagate(False)

        header = tk.Frame(card, bg=C_TEAL_LIGHT)
        header.pack(fill="x")
        tk.Label(header, text=f"{entry['weekday'][:2]} {entry['date'][:5]}",
                 font=("Segoe UI", 9, "bold"), bg=C_TEAL_LIGHT,
                 fg=C_TEAL_DARK).pack(padx=4, pady=2)

        haupt   = entry.get("menu", {}).get("haupt", [])
        preview = haupt[0] if haupt else "–"
        if len(preview) > 14:
            preview = preview[:13] + "…"
        tk.Label(card, text=preview, font=FONT_TINY,
                 bg=C_WHITE, fg=C_GRAY, wraplength=100).pack(padx=4, pady=(2, 4))

        def load(_e=None, en=entry):
            try:
                d = datetime.strptime(en["date"], "%d.%m.%Y")
                self._set_date(d)
            except Exception:
                pass
            self._menu_in_felder(en.get("menu", {}))
            self._set_status(f"Menü vom {en['date']} geladen.", C_TEAL)
            self._schedule_render()

        for w in (card, header, *card.winfo_children(), *header.winfo_children()):
            w.bind("<Button-1>", load)

    # ── Hilfen ────────────────────────────────────────────────────────────────

    @staticmethod
    def _hex_to_rgb(hex_color: str) -> tuple:
        h = hex_color.lstrip("#")
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

    # ── Start ─────────────────────────────────────────────────────────────────

    def run(self):
        self.root.update_idletasks()
        w = 1240
        h = 920
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{max(0,(sw-w)//2)}+{max(0,(sh-h)//2)}")
        self.root.mainloop()


# ══════════════════════════════════════════════════════════════════════════════
# Einstiegspunkt
# ══════════════════════════════════════════════════════════════════════════════

def _write_crash_log(err: BaseException) -> Path:
    """Schreibt einen Crash-Log mit Traceback neben das Skript."""
    import traceback
    log_path = Path(__file__).parent / "tagesbuffet_crash.log"
    try:
        with open(log_path, "w", encoding="utf-8") as f:
            f.write(f"Tagesbuffet-GUI Crash – {datetime.now().isoformat()}\n")
            f.write("=" * 60 + "\n")
            f.write(f"Python: {sys.version}\n")
            f.write(f"Platform: {sys.platform}\n")
            f.write(f"Skript: {__file__}\n")
            f.write(f"CWD: {os.getcwd()}\n")
            f.write(f"PIL_OK: {PIL_OK}, OPENPYXL_OK: {OPENPYXL_OK}, "
                    f"ANTHROPIC_OK: {ANTHROPIC_OK}, GENERATOR_OK: {GENERATOR_OK}\n")
            f.write("=" * 60 + "\n\n")
            traceback.print_exception(type(err), err, err.__traceback__, file=f)
    except Exception:
        pass
    return log_path


def main():
    try:
        if not GENERATOR_OK:
            try:
                root = tk.Tk(); root.withdraw()
                messagebox.showerror(
                    "Tagesbuffet – Fehler",
                    "tagesbuffet_generator.py nicht gefunden.\n\n"
                    "Bitte beide Dateien im selben Ordner ablegen:\n"
                    "  tagesbuffet_generator.py\n"
                    "  tagesbuffet_gui.py"
                )
                root.destroy()
            except Exception:
                pass
            sys.exit(1)

        app = TagesbuffetGUI()
        app.run()

    except BaseException as e:
        log_path = _write_crash_log(e)
        try:
            root = tk.Tk(); root.withdraw()
            messagebox.showerror(
                "Tagesbuffet – Crash",
                f"Die GUI ist abgestürzt.\n\n"
                f"Fehler: {type(e).__name__}: {e}\n\n"
                f"Details siehe:\n{log_path}"
            )
            root.destroy()
        except Exception:
            pass
        raise


if __name__ == "__main__":
    main()
