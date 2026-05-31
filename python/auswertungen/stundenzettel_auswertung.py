"""Stundenzettel-Auswertung für das Hotel Aquarius.

Dieses Modul nimmt die aus den Mitarbeiter-Stundenzetteln abgelesenen
Tagesdaten entgegen, berechnet pro Tag die Netto-Arbeitszeit (abzüglich Pause)
und schreibt das Ergebnis in eine gemeinsame Excel-Datei (ein Tabellenblatt je
Mitarbeiter plus ein Übersichtsblatt).

Berechnungslogik (mit dem Inhaber abgestimmt):
    Netto-Arbeitszeit = (bis - von) - Pause
    - Pause wird so angesetzt, wie auf dem Zettel eingetragen.
    - Ist KEINE Pause eingetragen, werden pauschal 30 Minuten abgezogen.
    - Ausgabe der Zeiten im Format Stunden:Minuten (z. B. 7:30).

Verwendung:
    Die Tagesdaten werden als Python-Strukturen übergeben (siehe
    ``MitarbeiterMonat``). Anschließend erzeugt bzw. aktualisiert
    ``schreibe_auswertung()`` die Excel-Datei. Ein konkretes Beispiel mit
    echten Daten findet sich in ``daten_mai_2026.py``.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

# --- Konstanten -----------------------------------------------------------

#: Standard-Pause in Minuten, wenn auf dem Zettel keine Pause eingetragen ist.
STANDARD_PAUSE_MIN: int = 30

#: Deutsche Wochentags-Kürzel (Montag = 0 ... Sonntag = 6).
WOCHENTAGE: tuple[str, ...] = ("Mo", "Di", "Mi", "Do", "Fr", "Sa", "So")

#: Deutsche Monatsnamen (Index 1 = Januar).
MONATSNAMEN: tuple[str, ...] = (
    "",
    "Januar", "Februar", "März", "April", "Mai", "Juni",
    "Juli", "August", "September", "Oktober", "November", "Dezember",
)

#: Pfad zur gemeinsamen Excel-Datei.
EXCEL_PFAD: Path = Path(__file__).parent / "stundenauswertung.xlsx"

# --- Styling --------------------------------------------------------------

_FARBE_HOTEL = "1F6F8B"          # dunkles Aquarius-Blau für Überschriften
_FARBE_KOPF = "BDE0E8"           # helles Blau für Tabellenkopf
_FARBE_WOCHENENDE = "FCE8B2"     # sanftes Gelb für Wochenenden
_FARBE_FREI = "F1F3F4"           # hellgrau für freie Tage
_FARBE_SUMME = "D9EAD3"          # helles Grün für Summenzeile

_DUENN = Side(style="thin", color="9AA0A6")
_RAHMEN = Border(left=_DUENN, right=_DUENN, top=_DUENN, bottom=_DUENN)


# --- Datenmodell ----------------------------------------------------------


@dataclass
class Tageseintrag:
    """Ein einzelner Arbeitstag eines Mitarbeiters.

    Attributes:
        tag: Tag im Monat (1-31).
        von: Arbeitsbeginn als "HH:MM" oder None, falls kein Eintrag (frei/krank).
        bis: Arbeitsende als "HH:MM" oder None.
        pause_min: Pause in Minuten. None bedeutet "keine Pause eingetragen"
            -> es werden STANDARD_PAUSE_MIN angesetzt. 0 bedeutet ausdrücklich
            keine Pause.
        bemerkung: Optionaler Hinweis (z. B. "krank", "Urlaub", "frei").
    """

    tag: int
    von: str | None = None
    bis: str | None = None
    pause_min: int | None = None
    bemerkung: str = ""

    def hat_arbeitszeit(self) -> bool:
        """Gibt True zurück, wenn ein gültiges Von/Bis-Paar vorliegt."""
        return bool(self.von and self.bis)

    def angesetzte_pause(self) -> int:
        """Tatsächlich abzuziehende Pause in Minuten.

        Returns:
            Eingetragene Pause oder STANDARD_PAUSE_MIN, falls None.
        """
        if not self.hat_arbeitszeit():
            return 0
        if self.pause_min is None:
            return STANDARD_PAUSE_MIN
        return self.pause_min

    def netto_minuten(self) -> int:
        """Berechnet die Netto-Arbeitszeit des Tages in Minuten.

        Returns:
            (bis - von) - Pause, in Minuten. 0, wenn kein Von/Bis vorliegt.
        """
        if not self.hat_arbeitszeit():
            return 0
        brutto = _zeit_in_minuten(self.bis) - _zeit_in_minuten(self.von)
        if brutto < 0:  # über Mitternacht hinaus
            brutto += 24 * 60
        return max(brutto - self.angesetzte_pause(), 0)


@dataclass
class MitarbeiterMonat:
    """Die Stundendaten eines Mitarbeiters für einen Monat.

    Attributes:
        name: Name des Mitarbeiters (wird zum Blattnamen).
        jahr: Jahr, z. B. 2026.
        monat: Monat als Zahl (1-12).
        eintraege: Liste der Tageseinträge.
    """

    name: str
    jahr: int
    monat: int
    eintraege: list[Tageseintrag] = field(default_factory=list)

    def gesamt_minuten(self) -> int:
        """Summe der Netto-Arbeitszeit aller Tage in Minuten."""
        return sum(e.netto_minuten() for e in self.eintraege)

    def arbeitstage(self) -> int:
        """Anzahl der Tage mit tatsächlicher Arbeitszeit."""
        return sum(1 for e in self.eintraege if e.netto_minuten() > 0)


# --- Hilfsfunktionen ------------------------------------------------------


def _zeit_in_minuten(zeit: str) -> int:
    """Wandelt "HH:MM" in Minuten seit Mitternacht um."""
    stunde, minute = zeit.strip().split(":")
    return int(stunde) * 60 + int(minute)


def minuten_als_hhmm(minuten: int) -> str:
    """Formatiert Minuten als "H:MM" (Stunden ohne führende Null).

    Beispiel:
        450 -> "7:30"
    """
    return f"{minuten // 60}:{minuten % 60:02d}"


# --- Excel-Erzeugung ------------------------------------------------------


def _setze_spaltenbreiten(ws: Worksheet, breiten: dict[str, int]) -> None:
    """Setzt die Spaltenbreiten anhand eines Spaltenbuchstaben-Mappings."""
    for spalte, breite in breiten.items():
        ws.column_dimensions[spalte].width = breite


def _schreibe_mitarbeiterblatt(wb: Workbook, daten: MitarbeiterMonat) -> None:
    """Erzeugt bzw. ersetzt das Tabellenblatt eines Mitarbeiters."""
    # Vorhandenes Blatt entfernen, damit Aktualisierungen sauber sind.
    if daten.name in wb.sheetnames:
        del wb[daten.name]
    ws = wb.create_sheet(title=daten.name)

    monatsname = MONATSNAMEN[daten.monat]

    # Kopfbereich
    ws["A1"] = "Hotel Aquarius – Stundenauswertung"
    ws["A1"].font = Font(bold=True, size=14, color=_FARBE_HOTEL)
    ws["A2"] = f"Mitarbeiter: {daten.name}"
    ws["A3"] = f"Monat: {monatsname} {daten.jahr}"
    for zelle in ("A2", "A3"):
        ws[zelle].font = Font(bold=True, size=11)

    # Tabellenkopf
    kopf = ["Datum", "Wochentag", "von", "bis", "Pause", "Stunden"]
    kopf_zeile = 5
    for spalte, titel in enumerate(kopf, start=1):
        zelle = ws.cell(row=kopf_zeile, column=spalte, value=titel)
        zelle.font = Font(bold=True)
        zelle.fill = PatternFill("solid", fgColor=_FARBE_KOPF)
        zelle.alignment = Alignment(horizontal="center")
        zelle.border = _RAHMEN

    # Datenzeilen
    zeile = kopf_zeile + 1
    for eintrag in sorted(daten.eintraege, key=lambda e: e.tag):
        datum = dt.date(daten.jahr, daten.monat, eintrag.tag)
        ist_wochenende = datum.weekday() >= 5
        arbeitet = eintrag.hat_arbeitszeit()

        werte = [
            datum.strftime("%d.%m.%Y"),
            WOCHENTAGE[datum.weekday()],
            eintrag.von or "",
            eintrag.bis or "",
            minuten_als_hhmm(eintrag.angesetzte_pause()) if arbeitet else "",
            minuten_als_hhmm(eintrag.netto_minuten()) if arbeitet else "",
        ]
        # Bei freien Tagen ggf. Bemerkung in der Stunden-Spalte zeigen.
        if not arbeitet and eintrag.bemerkung:
            werte[5] = eintrag.bemerkung

        for spalte, wert in enumerate(werte, start=1):
            zelle = ws.cell(row=zeile, column=spalte, value=wert)
            zelle.border = _RAHMEN
            zelle.alignment = Alignment(horizontal="center")
            if ist_wochenende:
                zelle.fill = PatternFill("solid", fgColor=_FARBE_WOCHENENDE)
            elif not arbeitet:
                zelle.fill = PatternFill("solid", fgColor=_FARBE_FREI)
        zeile += 1

    # Summenzeile
    summe_zeile = zeile
    ws.cell(row=summe_zeile, column=1, value="Summe").font = Font(bold=True)
    ws.cell(row=summe_zeile, column=2, value=f"{daten.arbeitstage()} Tage")
    gesamt = ws.cell(
        row=summe_zeile, column=6, value=minuten_als_hhmm(daten.gesamt_minuten())
    )
    gesamt.font = Font(bold=True)
    for spalte in range(1, 7):
        zelle = ws.cell(row=summe_zeile, column=spalte)
        zelle.fill = PatternFill("solid", fgColor=_FARBE_SUMME)
        zelle.border = _RAHMEN
        zelle.alignment = Alignment(horizontal="center")

    _setze_spaltenbreiten(
        ws, {"A": 12, "B": 11, "C": 8, "D": 8, "E": 8, "F": 11}
    )


def _aktualisiere_uebersicht(wb: Workbook, alle: list[MitarbeiterMonat]) -> None:
    """Erzeugt bzw. aktualisiert das Übersichtsblatt über alle Mitarbeiter."""
    titel = "Übersicht"
    if titel in wb.sheetnames:
        del wb[titel]
    ws = wb.create_sheet(title=titel, index=0)

    ws["A1"] = "Hotel Aquarius – Stundenübersicht"
    ws["A1"].font = Font(bold=True, size=14, color=_FARBE_HOTEL)

    kopf = ["Mitarbeiter", "Monat", "Arbeitstage", "Gesamtstunden"]
    for spalte, t in enumerate(kopf, start=1):
        zelle = ws.cell(row=3, column=spalte, value=t)
        zelle.font = Font(bold=True)
        zelle.fill = PatternFill("solid", fgColor=_FARBE_KOPF)
        zelle.alignment = Alignment(horizontal="center")
        zelle.border = _RAHMEN

    zeile = 4
    for daten in sorted(alle, key=lambda d: (d.jahr, d.monat, d.name)):
        werte = [
            daten.name,
            f"{MONATSNAMEN[daten.monat]} {daten.jahr}",
            daten.arbeitstage(),
            minuten_als_hhmm(daten.gesamt_minuten()),
        ]
        for spalte, wert in enumerate(werte, start=1):
            zelle = ws.cell(row=zeile, column=spalte, value=wert)
            zelle.border = _RAHMEN
            zelle.alignment = Alignment(horizontal="center")
        zeile += 1

    _setze_spaltenbreiten(ws, {"A": 18, "B": 16, "C": 13, "D": 14})


def schreibe_auswertung(
    mitarbeiter: list[MitarbeiterMonat], pfad: Path = EXCEL_PFAD
) -> Path:
    """Schreibt die Auswertung aller Mitarbeiter in die Excel-Datei.

    Vorhandene Blätter der übergebenen Mitarbeiter werden aktualisiert,
    bereits gespeicherte andere Mitarbeiter bleiben erhalten.

    Args:
        mitarbeiter: Liste der auszuwertenden Mitarbeiter-Monate.
        pfad: Zielpfad der Excel-Datei.

    Returns:
        Der Pfad der geschriebenen Datei.
    """
    if pfad.exists():
        wb = load_workbook(pfad)
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            del wb["Sheet"]
    else:
        wb = Workbook()
        del wb["Sheet"]

    for daten in mitarbeiter:
        _schreibe_mitarbeiterblatt(wb, daten)

    _aktualisiere_uebersicht(wb, mitarbeiter)

    wb.save(pfad)
    return pfad
