#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Hotel Aquarius - Verfügbarkeits- & Preis-Tool
==============================================
Liest automatisch alle Daten aus kurzurlaub.de/rezeption aus:
  - Zimmerkategorien & Anzahl
  - Belegungskalender (Sperrdaten)
  - Arrangements mit Saisonpreisen
Zeigt alles in einer interaktiven GUI an und kann als Excel exportiert werden.

Voraussetzungen:
  pip install requests beautifulsoup4 openpyxl

Nutzung:
  1. Im Browser auf kurzurlaub.de/rezeption einloggen
  2. Die komplette URL aus der Adresszeile kopieren
  3. Im Tool einfügen und "Verbinden" klicken
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import re
import calendar
from datetime import datetime, date, timedelta
from urllib.parse import urlparse, parse_qs, urlencode

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    requests = None
    BeautifulSoup = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─── Farbschema ───
COLORS = {
    "bg": "#f0f4f8", "primary": "#1a5276", "primary_light": "#2980b9",
    "accent": "#e67e22", "green": "#27ae60", "green_light": "#d5f5e3",
    "red": "#c0392b", "red_light": "#fadbd8", "blue_light": "#d6eaf8",
    "yellow_light": "#fef9e7", "white": "#ffffff", "gray": "#ecf0f1",
    "gray_dark": "#7f8c8d", "text": "#2c3e50",
}
ROOM_COLORS = ["#3498db", "#2ecc71", "#9b59b6", "#e67e22", "#e74c3c",
               "#1abc9c", "#f1c40f", "#e91e63", "#00bcd4", "#8bc34a"]
MONTHS_DE = ["Januar", "Februar", "März", "April", "Mai", "Juni",
             "Juli", "August", "September", "Oktober", "November", "Dezember"]
DAYS_SHORT = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]


# ─────────────────────────────────────────────────
# Scraper
# ─────────────────────────────────────────────────
class KurzurlaubScraper:
    BASE = "https://www.kurzurlaub.de/rezeption"

    def __init__(self):
        self.session_token = ""
        self.o_id = ""
        self.http = requests.Session()
        self.http.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 Chrome/120 Safari/537.36"
        })
        self.rooms = []       # [{id, name, size, count, color}]
        self.blocked = {}     # {room_id: set(date_str)}
        self.arrangements = []
        self.seasons = []     # [{name, ranges:[(start,end),...], prices:{room_id: price}}]
        self.connected = False

    def _url(self, path, **extra):
        params = {"session": self.session_token, "o_id": self.o_id}
        params.update(extra)
        return f"{self.BASE}/{path}?{urlencode(params)}"

    # ── Verbinden ──
    def connect(self, full_url):
        parsed = urlparse(full_url)
        params = parse_qs(parsed.query)
        self.session_token = params.get("session", [""])[0]
        self.o_id = params.get("o_id", ["7348"])[0]

        if not self.session_token:
            raise ValueError(
                "Kein 'session'-Parameter in der URL gefunden.\n"
                "Bitte die komplette URL aus dem Browser kopieren.")

        r = self.http.get(self._url("startseite.html"), timeout=15)
        text = r.text.lower()
        if "log out" in text or "dashboard" in text or "hotel aquarius" in text:
            self.connected = True
            return True
        if "login" in text or "passwort" in text or "anmelden" in text:
            raise ConnectionError(
                "Session abgelaufen!\n"
                "Bitte im Browser neu einloggen und\n"
                "die aktuelle URL erneut kopieren.")
        self.connected = True
        return True

    # ── Zimmerkategorien ──
    def load_rooms(self, callback=None):
        url = self._url("hotel/sperrdaten/einfach.html")
        r = self.http.get(url, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")
        self.rooms = []

        # Radiobuttons: name="form[avail][room]"
        radios = soup.find_all("input", {"name": "form[avail][room]"})
        for radio in radios:
            val = radio.get("value", "").strip()
            if not val or val == "0":
                continue

            # Label-Text steht im Eltern-Element
            parent = radio.find_parent("label") or radio.parent
            if parent:
                label_text = parent.get_text(strip=True)
            else:
                label_text = f"Zimmer {val}"

            # Entferne Radio-Markierung und ID-Angabe
            label_text = re.sub(r"\(ID\s*\d+\)", "", label_text).strip()

            count, name, size = self._parse_room_name(label_text)
            idx = len(self.rooms)
            self.rooms.append({
                "id": int(val), "name": name, "size": size,
                "count": count,
                "color": ROOM_COLORS[idx % len(ROOM_COLORS)]
            })

        if callback:
            callback(f"{len(self.rooms)} Zimmerkategorien gefunden")
        return self.rooms

    def _parse_room_name(self, text):
        count = 1
        m = re.match(r"(\d+)\s+(.+)", text)
        if m:
            count = int(m.group(1))
            text = m.group(2)
        size = ""
        m2 = re.search(r"\(([^)]*m²[^)]*)\)", text)
        if m2:
            size = m2.group(1)
            text = text[:m2.start()].strip()
        return count, text.strip(), size

    # ── Belegungskalender ──
    def load_calendar(self, year=None, callback=None):
        if year is None:
            year = date.today().year
        self.blocked = {}

        for i, room in enumerate(self.rooms):
            rid = room["id"]
            self.blocked[rid] = set()
            if callback:
                callback(f"Lade Kalender: {room['name']}... ({i+1}/{len(self.rooms)})")

            # Die Seite zeigt alle Monate auf einmal per GET mit zimmer=ID
            url = self._url("hotel/sperrdaten/einfach.html", zimmer=rid)
            try:
                r = self.http.get(url, timeout=20)
                soup = BeautifulSoup(r.text, "html.parser")
                self._parse_calendar_hidden_inputs(soup, rid)
            except Exception as e:
                if callback:
                    callback(f"Fehler bei {room['name']}: {e}")

        if callback:
            total = sum(len(v) for v in self.blocked.values())
            callback(f"Kalender geladen: {total} Sperrtage insgesamt")

    def _parse_calendar_hidden_inputs(self, soup, room_id):
        """Lese Sperrdaten aus hidden inputs: form[avail][count][0][YYYY-MM-DD]
        Wert 0 = gesperrt, >0 = verfügbar, -1 = vergangen"""
        hiddens = soup.find_all("input", {"type": "hidden"})
        for inp in hiddens:
            name = inp.get("name", "")
            m = re.search(r"form\[avail\]\[count\]\[\d+\]\[(\d{4}-\d{2}-\d{2})\]", name)
            if m:
                date_str = m.group(1)
                val = inp.get("value", "")
                try:
                    if int(val) == 0:
                        self.blocked[room_id].add(date_str)
                except ValueError:
                    pass

    # ── Arrangements & Preise ──
    def load_arrangements(self, callback=None):
        self.arrangements = []

        url = self._url("arrangements/alle.html")
        r = self.http.get(url, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")

        # Finde offer_ids aus "Bearbeiten"-Links
        offer_ids = []
        seen = set()
        for a in soup.find_all("a", href=True):
            href = a["href"]
            m = re.search(r"offer_id=(\d+)", href)
            if m:
                oid = m.group(1)
                if oid != "0" and oid not in seen:
                    seen.add(oid)
                    offer_ids.append(oid)

        if callback:
            callback(f"{len(offer_ids)} Arrangements gefunden")

        for idx, oid in enumerate(offer_ids):
            if callback:
                callback(f"Lade Arrangement {idx+1}/{len(offer_ids)}...")
            try:
                self._load_arrangement_detail(oid)
            except Exception as e:
                if callback:
                    callback(f"Fehler bei Arrangement {oid}: {e}")

        if callback:
            callback(f"{len(self.arrangements)} Arrangements mit Preisen geladen")

    def _load_arrangement_detail(self, offer_id):
        url = self._url("arrangements/bearbeiten.html", offer_id=offer_id)
        r = self.http.get(url, timeout=20)
        soup = BeautifulSoup(r.text, "html.parser")

        # Arrangement-Name aus der Überschrift
        h1 = soup.find("h1")
        arr_name = ""
        if h1:
            arr_name = h1.get_text(strip=True)
            # Bereinigen: "...bearbeiten" entfernen
            arr_name = re.sub(r'[„""]', '', arr_name)
            arr_name = arr_name.replace(" bearbeiten", "").strip()

        # Übernachtungen aus Select
        nights = ""
        select = soup.find("select", {"name": re.compile("nights|uebernachtung", re.I)})
        if not select:
            for sel in soup.find_all("select"):
                for opt in sel.find_all("option", selected=True):
                    t = opt.get_text(strip=True)
                    if "Nächte" in t or "Tage" in t:
                        nights = t
                        break
                if nights:
                    break

        # Kurzname generieren
        m = re.search(r"(\d+)\s*Tage", arr_name)
        m2 = re.search(r"(\d+)\s*Nächte?\s*/\s*(\d+)\s*Tage", nights) if nights else None
        short_name = arr_name[:40]
        if m:
            tage = int(m.group(1))
            naechte = tage - 1
            short_name = f"{tage}T/{naechte}N"
        elif m2:
            short_name = f"{m2.group(2)}T/{m2.group(1)}N"

        # ── Basispreise aus Input-Feldern ──
        basis_prices = {}
        for inp in soup.find_all("input", {"name": re.compile(r"offer_room_price\[\d+\]")}):
            m = re.search(r"offer_room_price\[(\d+)\]", inp["name"])
            if m:
                rid = int(m.group(1))
                try:
                    price = float(inp.get("value", "0").replace(",", "."))
                    if price > 0:
                        basis_prices[rid] = price
                except ValueError:
                    pass

        # ── Saisonpreise aus dem Text-Block ──
        season_prices = {}  # {season_name: {room_id: price}}
        season_ranges = {}  # {season_name: [(start, end), ...]}

        page_text = soup.get_text()
        saison_idx = page_text.find("Folgende Saisonpreise")
        if saison_idx >= 0:
            saison_text = page_text[saison_idx:]

            # Finde Saison-Blöcke: z.B. "Hauptsaison (22.12.2025 - 02.01.2026,...)"
            # gefolgt von Zimmer-Preisen
            pattern = re.compile(
                r"(Haupt|Neben|Sonder)saison\s*\(([^)]+)\)",
                re.IGNORECASE
            )
            matches = list(pattern.finditer(saison_text))

            for mi, match in enumerate(matches):
                season_name = match.group(1).lower() + "saison"
                dates_str = match.group(2)

                # Parse Datumsbereiche
                ranges = []
                date_pairs = re.findall(
                    r"(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})",
                    dates_str)
                for s, e in date_pairs:
                    try:
                        start = datetime.strptime(s, "%d.%m.%Y").date()
                        end = datetime.strptime(e, "%d.%m.%Y").date()
                        ranges.append((start, end))
                    except ValueError:
                        pass
                season_ranges[season_name] = ranges

                # Preise: Text zwischen diesem und nächstem Saison-Match
                end_pos = matches[mi + 1].start() if mi + 1 < len(matches) \
                    else len(saison_text)
                block = saison_text[match.end():end_pos]

                prices = {}
                # Robusterer Ansatz: Block nach "(ID XXXXX)" aufteilen
                # und im jeweiligen Teilstück den ersten Preis finden
                id_splits = re.split(r"\(ID\s*(\d+)\)", block)
                # id_splits: [text_before, id1, text_after_id1, id2, text_after_id2, ...]
                for si in range(1, len(id_splits) - 1, 2):
                    rid = int(id_splits[si])
                    chunk = id_splits[si + 1]
                    # Suche den ersten Preis im Chunk (z.B. "159.00€" oder "159,00 €")
                    pm = re.search(r"(\d+[\.,]\d{2})\s*€", chunk)
                    if pm:
                        price = float(pm.group(1).replace(",", "."))
                        if price > 0:
                            prices[rid] = price

                season_prices[season_name] = prices

        # ── Ergebnis zusammenbauen ──
        arr = {
            "id": offer_id,
            "name": arr_name,
            "short_name": short_name,
            "basis_prices": basis_prices,
            "season_prices": season_prices,
            "season_ranges": season_ranges,
        }

        # Nur hinzufügen wenn Preise vorhanden
        if basis_prices:
            self.arrangements.append(arr)

            # Seasons global speichern (einmal reicht)
            if season_ranges and not self.seasons:
                for sname, ranges in season_ranges.items():
                    self.seasons.append({
                        "name": sname,
                        "ranges": ranges,
                    })

    # ── Hilfsmethoden ──
    def is_blocked(self, room_id, d):
        return d.isoformat() in self.blocked.get(room_id, set())

    def get_available_count(self, room_id, d):
        room = next((r for r in self.rooms if r["id"] == room_id), None)
        if not room:
            return 0
        return 0 if self.is_blocked(room_id, d) else room["count"]

    def get_total_available(self, d, room_filter=None):
        total = avail = 0
        for room in self.rooms:
            if room_filter and room["id"] not in room_filter:
                continue
            total += room["count"]
            if not self.is_blocked(room["id"], d):
                avail += room["count"]
        return avail, total

    def get_season(self, d):
        for s in self.seasons:
            for start, end in s["ranges"]:
                if start <= d <= end:
                    return s["name"]
        return "basis"

    def get_price(self, room_id, arr_idx, d):
        if arr_idx >= len(self.arrangements):
            return None
        arr = self.arrangements[arr_idx]
        season = self.get_season(d)
        if season != "basis" and season in arr["season_prices"]:
            p = arr["season_prices"][season].get(room_id)
            if p:
                return p
        return arr["basis_prices"].get(room_id)

    def get_price_for_season(self, room_id, arr_idx, season_name):
        if arr_idx >= len(self.arrangements):
            return None
        arr = self.arrangements[arr_idx]
        if season_name in arr["season_prices"]:
            return arr["season_prices"][season_name].get(room_id)
        return arr["basis_prices"].get(room_id)


# ─────────────────────────────────────────────────
# GUI
# ─────────────────────────────────────────────────
class HotelApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Hotel Aquarius - Verfügbarkeits- & Preis-Tool")
        self.geometry("1200x800")
        self.configure(bg=COLORS["bg"])
        self.minsize(900, 600)

        self.scraper = KurzurlaubScraper()
        self.active_rooms = set()
        self.current_year = date.today().year
        self.current_month = date.today().month
        self._tooltip_win = None

        self._build_styles()
        self._build_ui()
        self._show_login()

    def _build_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Tab.TNotebook.Tab",
                         font=("Segoe UI", 10, "bold"), padding=[15, 8])

    def _build_ui(self):
        header = tk.Frame(self, bg=COLORS["primary"], height=60)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        hbox = tk.Frame(header, bg=COLORS["primary"])
        hbox.pack(fill=tk.BOTH, expand=True, padx=20, pady=8)

        tk.Label(hbox, text="Hotel Aquarius - Norddeich",
                 bg=COLORS["primary"], fg="white",
                 font=("Segoe UI", 16, "bold")).pack(side=tk.LEFT)

        self.btn_export = tk.Button(
            hbox, text="Excel Export", bg=COLORS["green"], fg="white",
            font=("Segoe UI", 10, "bold"), relief="flat", padx=15, pady=4,
            command=self._export_excel, cursor="hand2")
        self.btn_export.pack(side=tk.RIGHT, padx=5)

        self.btn_refresh = tk.Button(
            hbox, text="Daten aktualisieren", bg=COLORS["accent"], fg="white",
            font=("Segoe UI", 10, "bold"), relief="flat", padx=15, pady=4,
            command=self._refresh_data, cursor="hand2")
        self.btn_refresh.pack(side=tk.RIGHT, padx=5)

        self.lbl_header = tk.Label(hbox, text="Nicht verbunden",
                                    bg=COLORS["primary"], fg="#f39c12",
                                    font=("Segoe UI", 9))
        self.lbl_header.pack(side=tk.RIGHT, padx=15)

        self.main_frame = tk.Frame(self, bg=COLORS["bg"])
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        self.status_var = tk.StringVar(value="Bereit")
        tk.Label(self, textvariable=self.status_var, bg=COLORS["gray"],
                 fg=COLORS["text"], font=("Segoe UI", 9),
                 anchor="w", padx=10).pack(fill=tk.X, side=tk.BOTTOM)

        self.progress = ttk.Progressbar(self, mode="indeterminate")

    # ── Login ──
    def _show_login(self):
        for w in self.main_frame.winfo_children():
            w.destroy()

        center = tk.Frame(self.main_frame, bg=COLORS["bg"])
        center.place(relx=0.5, rely=0.45, anchor="center")

        card = tk.Frame(center, bg=COLORS["white"], padx=40, pady=30,
                        highlightbackground="#ddd", highlightthickness=1)
        card.pack()

        tk.Label(card, text="Verbindung herstellen",
                 bg=COLORS["white"], fg=COLORS["primary"],
                 font=("Segoe UI", 18, "bold")).pack(pady=(0, 5))
        tk.Label(card, text="kurzurlaub.de / Rezeption",
                 bg=COLORS["white"], fg=COLORS["gray_dark"],
                 font=("Segoe UI", 10)).pack(pady=(0, 20))

        info = tk.Frame(card, bg=COLORS["blue_light"], padx=15, pady=12)
        info.pack(fill=tk.X, pady=(0, 15))
        tk.Label(info, text="So geht's:", bg=COLORS["blue_light"],
                 fg=COLORS["primary"],
                 font=("Segoe UI", 10, "bold")).pack(anchor="w")
        for step in [
            "1. Im Browser auf kurzurlaub.de/rezeption einloggen",
            "2. Die komplette URL aus der Adresszeile kopieren",
            "   (enthält session=... und o_id=...)",
            "3. Hier einfügen und 'Verbinden' klicken"
        ]:
            tk.Label(info, text=step, bg=COLORS["blue_light"],
                     fg=COLORS["text"], font=("Segoe UI", 9)).pack(anchor="w")

        tk.Label(card, text="URL aus dem Browser:",
                 bg=COLORS["white"], fg=COLORS["text"],
                 font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(5, 0))

        self.entry_url = tk.Entry(card, font=("Segoe UI", 11), width=65)
        self.entry_url.pack(fill=tk.X, pady=(5, 3), ipady=6)

        tk.Label(card,
                 text="z.B.: https://www.kurzurlaub.de/rezeption/startseite.html?session=abc&o_id=7348",
                 bg=COLORS["white"], fg=COLORS["gray_dark"],
                 font=("Segoe UI", 8, "italic")).pack(anchor="w", pady=(0, 15))

        self.btn_connect = tk.Button(
            card, text="Verbinden & Daten laden",
            bg=COLORS["primary_light"], fg="white",
            font=("Segoe UI", 12, "bold"), relief="flat",
            padx=20, pady=10, cursor="hand2", command=self._do_connect)
        self.btn_connect.pack(fill=tk.X)

        self.login_status = tk.Label(card, text="", bg=COLORS["white"],
                                      font=("Segoe UI", 9))
        self.login_status.pack(pady=(10, 0))

        # Auto-paste
        try:
            clip = self.clipboard_get()
            if "kurzurlaub.de" in clip and "session=" in clip:
                self.entry_url.insert(0, clip)
                self.login_status.config(text="URL aus Zwischenablage eingefügt!",
                                          fg=COLORS["green"])
        except Exception:
            pass

    def _do_connect(self):
        url = self.entry_url.get().strip()
        if not url:
            self.login_status.config(text="Bitte URL eingeben.",
                                      fg=COLORS["red"])
            return
        if "session=" not in url:
            self.login_status.config(
                text="Kein 'session'-Parameter gefunden. Bitte die vollständige URL kopieren.",
                fg=COLORS["red"])
            return

        self.btn_connect.config(state="disabled", text="Verbinde...")
        self.login_status.config(text="Verbinde mit kurzurlaub.de...",
                                  fg=COLORS["primary_light"])
        if not self.progress.winfo_ismapped():
            self.progress.pack(fill=tk.X, side=tk.BOTTOM)
        self.progress.start()

        def worker():
            try:
                self.scraper.connect(url)
                self.after(0, self._connect_ok)
            except Exception as e:
                self.after(0, lambda: self._connect_err(str(e)))

        threading.Thread(target=worker, daemon=True).start()

    def _connect_ok(self):
        self.login_status.config(text="Verbunden! Lade Daten...",
                                  fg=COLORS["green"])
        self.lbl_header.config(text="Verbunden", fg=COLORS["green"])
        self._load_all_data()

    def _connect_err(self, err):
        self.progress.stop()
        self.progress.pack_forget()
        self.login_status.config(text=f"Fehler: {err}", fg=COLORS["red"])
        self.btn_connect.config(state="normal", text="Verbinden & Daten laden")

    # ── Daten laden ──
    def _load_all_data(self):
        if not self.progress.winfo_ismapped():
            self.progress.pack(fill=tk.X, side=tk.BOTTOM)
        self.progress.start()

        def worker():
            try:
                cb = lambda m: self.after(0, lambda m=m: self.status_var.set(m))

                cb("Lade Zimmerkategorien...")
                self.scraper.load_rooms(callback=cb)

                cb("Lade Belegungskalender...")
                self.scraper.load_calendar(self.current_year, callback=cb)

                cb("Lade Arrangements & Preise...")
                self.scraper.load_arrangements(callback=cb)

                self.after(0, self._data_loaded)
            except Exception as e:
                self.after(0, lambda: self._data_error(str(e)))

        threading.Thread(target=worker, daemon=True).start()

    def _refresh_data(self):
        if not self.scraper.connected:
            messagebox.showwarning("Nicht verbunden", "Bitte zuerst verbinden.")
            return
        self._load_all_data()

    def _data_loaded(self):
        self.progress.stop()
        self.progress.pack_forget()
        self.active_rooms = set(r["id"] for r in self.scraper.rooms)

        nr = len(self.scraper.rooms)
        na = len(self.scraper.arrangements)
        tb = sum(len(v) for v in self.scraper.blocked.values())
        self.status_var.set(f"Geladen: {nr} Kategorien, {na} Arrangements, {tb} Sperrtage")
        self.lbl_header.config(
            text=f"Verbunden | {nr} Kategorien | {na} Arrangements",
            fg=COLORS["green"])
        self._build_main_view()

    def _data_error(self, err):
        self.progress.stop()
        self.progress.pack_forget()
        self.status_var.set(f"Fehler: {err}")
        messagebox.showerror("Fehler", f"Fehler beim Laden:\n{err}")

    # ── Hauptansicht ──
    def _build_main_view(self):
        for w in self.main_frame.winfo_children():
            w.destroy()

        # Stats
        sf = tk.Frame(self.main_frame, bg=COLORS["bg"])
        sf.pack(fill=tk.X, padx=15, pady=(10, 5))
        self._build_stats(sf)

        # Filter
        ff = tk.Frame(self.main_frame, bg=COLORS["white"],
                       highlightbackground="#e0e0e0", highlightthickness=1)
        ff.pack(fill=tk.X, padx=15, pady=5)
        self._build_filter(ff)

        # Tabs
        nb = ttk.Notebook(self.main_frame, style="Tab.TNotebook")
        nb.pack(fill=tk.BOTH, expand=True, padx=15, pady=(5, 10))

        self.tab_cal = tk.Frame(nb, bg=COLORS["white"])
        nb.add(self.tab_cal, text="  Monatsansicht  ")
        self.tab_prices = tk.Frame(nb, bg=COLORS["white"])
        nb.add(self.tab_prices, text="  Preisübersicht  ")
        self.tab_detail = tk.Frame(nb, bg=COLORS["white"])
        nb.add(self.tab_detail, text="  Detailansicht  ")

        self._render_calendar()
        self._render_prices()
        self._render_details()

    def _build_stats(self, parent):
        today = date.today()
        end = date(self.current_year, 12, 31)
        tot = av = 0
        d = today
        while d <= end:
            for r in self.scraper.rooms:
                if r["id"] in self.active_rooms:
                    tot += 1
                    if not self.scraper.is_blocked(r["id"], d):
                        av += 1
            d += timedelta(days=1)
        bl = tot - av
        pct = round(av / max(tot, 1) * 100)
        tz = sum(r["count"] for r in self.scraper.rooms if r["id"] in self.active_rooms)

        for label, val, det, col in [
            ("Kategorien", len([r for r in self.scraper.rooms if r["id"] in self.active_rooms]),
             f"{tz} Zimmer", COLORS["primary_light"]),
            ("Verfügbar", av, f"{pct}%", COLORS["green"]),
            ("Gesperrt", bl, f"{100-pct}%", COLORS["red"]),
            ("Arrangements", len(self.scraper.arrangements),
             f"bis Dez {self.current_year}", COLORS["accent"]),
        ]:
            c = tk.Frame(parent, bg=COLORS["white"], padx=15, pady=10,
                          highlightbackground="#e0e0e0", highlightthickness=1)
            c.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=3)
            tk.Frame(c, bg=col, width=4).pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
            inner = tk.Frame(c, bg=COLORS["white"])
            inner.pack(side=tk.LEFT)
            tk.Label(inner, text=label.upper(), bg=COLORS["white"],
                     fg=COLORS["gray_dark"], font=("Segoe UI", 8, "bold")).pack(anchor="w")
            tk.Label(inner, text=str(val), bg=COLORS["white"], fg=col,
                     font=("Segoe UI", 20, "bold")).pack(anchor="w")
            tk.Label(inner, text=det, bg=COLORS["white"],
                     fg=COLORS["gray_dark"], font=("Segoe UI", 8)).pack(anchor="w")

    def _build_filter(self, parent):
        inner = tk.Frame(parent, bg=COLORS["white"], padx=10, pady=8)
        inner.pack(fill=tk.X)
        tk.Label(inner, text="Kategorien:", bg=COLORS["white"],
                 fg=COLORS["primary"], font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        self.room_vars = {}
        for r in self.scraper.rooms:
            v = tk.BooleanVar(value=r["id"] in self.active_rooms)
            self.room_vars[r["id"]] = v
            tk.Checkbutton(inner, text=f"{r['count']}x {r['name']}", variable=v,
                           bg=COLORS["white"], fg=r["color"],
                           selectcolor=COLORS["white"],
                           activebackground=COLORS["white"],
                           font=("Segoe UI", 9, "bold"),
                           command=self._filter_changed).pack(side=tk.LEFT, padx=8)

    def _filter_changed(self):
        self.active_rooms = {rid for rid, v in self.room_vars.items() if v.get()}
        self._render_calendar()
        self._render_details()

    # ── Monatsansicht ──
    def _render_calendar(self):
        for w in self.tab_cal.winfo_children():
            w.destroy()

        nav = tk.Frame(self.tab_cal, bg=COLORS["white"])
        nav.pack(fill=tk.X, padx=10, pady=5)
        tk.Button(nav, text="<", command=self._prev_m,
                  bg=COLORS["primary_light"], fg="white",
                  font=("Segoe UI", 10, "bold"), relief="flat", padx=10,
                  cursor="hand2").pack(side=tk.LEFT)
        self.m_label = tk.Label(nav,
            text=f"{MONTHS_DE[self.current_month-1]} {self.current_year}",
            bg=COLORS["white"], fg=COLORS["primary"],
            font=("Segoe UI", 14, "bold"))
        self.m_label.pack(side=tk.LEFT, padx=15)
        tk.Button(nav, text=">", command=self._next_m,
                  bg=COLORS["primary_light"], fg="white",
                  font=("Segoe UI", 10, "bold"), relief="flat", padx=10,
                  cursor="hand2").pack(side=tk.LEFT)
        tk.Button(nav, text="Heute", command=self._today,
                  bg=COLORS["gray"], fg=COLORS["text"],
                  font=("Segoe UI", 9), relief="flat", padx=10,
                  cursor="hand2").pack(side=tk.LEFT, padx=10)

        canvas = tk.Canvas(self.tab_cal, bg=COLORS["white"], highlightthickness=0)
        sb = ttk.Scrollbar(self.tab_cal, orient="vertical", command=canvas.yview)
        sf = tk.Frame(canvas, bg=COLORS["white"])
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=sf, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.bind_all("<MouseWheel>",
                         lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        mf = tk.Frame(sf, bg=COLORS["white"])
        mf.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        for off in range(3):
            m = self.current_month + off
            y = self.current_year
            if m > 12:
                m -= 12; y += 1
            if y > self.current_year:
                break
            self._month_card(mf, y, m)

    def _month_card(self, parent, year, month):
        card = tk.Frame(parent, bg=COLORS["white"], padx=5, pady=5,
                         highlightbackground="#e0e0e0", highlightthickness=1)
        card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)

        hdr = tk.Frame(card, bg=COLORS["primary"], padx=10, pady=8)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text=f"{MONTHS_DE[month-1]} {year}",
                 bg=COLORS["primary"], fg="white",
                 font=("Segoe UI", 12, "bold")).pack()

        dh = tk.Frame(card, bg=COLORS["white"])
        dh.pack(fill=tk.X, pady=(5, 0))
        for d in DAYS_SHORT:
            fg = COLORS["red"] if d in ("Sa", "So") else COLORS["gray_dark"]
            tk.Label(dh, text=d, bg=COLORS["white"], fg=fg,
                     font=("Segoe UI", 8, "bold"), width=5).pack(
                         side=tk.LEFT, expand=True)

        for week in calendar.monthcalendar(year, month):
            rf = tk.Frame(card, bg=COLORS["white"])
            rf.pack(fill=tk.X)
            for dn in week:
                if dn == 0:
                    tk.Label(rf, text="", bg=COLORS["white"], width=5,
                             height=2).pack(side=tk.LEFT, expand=True)
                    continue
                d = date(year, month, dn)
                av, tot = self.scraper.get_total_available(d, self.active_rooms)
                if d < date.today():
                    bg, fg, sub = "#f0f0f0", "#ccc", ""
                elif av == 0:
                    bg, fg, sub = COLORS["red_light"], COLORS["red"], "0"
                elif av < tot:
                    bg, fg, sub = COLORS["yellow_light"], "#b7950b", str(av)
                else:
                    bg, fg, sub = COLORS["green_light"], COLORS["green"], str(av)

                cell = tk.Frame(rf, bg=bg, padx=2, pady=1)
                cell.pack(side=tk.LEFT, expand=True, fill=tk.BOTH, padx=1, pady=1)
                tk.Label(cell, text=str(dn), bg=bg, fg=fg,
                         font=("Segoe UI", 10, "bold")).pack()
                if sub:
                    tk.Label(cell, text=sub, bg=bg, fg=fg,
                             font=("Segoe UI", 7)).pack()
                for w in [cell] + list(cell.winfo_children()):
                    w.bind("<Enter>", lambda e, dt=d: self._tt_show(e, dt))
                    w.bind("<Leave>", lambda e: self._tt_hide())

    def _prev_m(self):
        self.current_month -= 1
        if self.current_month < 1: self.current_month = 12; self.current_year -= 1
        self._render_calendar()

    def _next_m(self):
        self.current_month += 1
        if self.current_month > 12: self.current_month = 1; self.current_year += 1
        self._render_calendar()

    def _today(self):
        self.current_month = date.today().month
        self.current_year = date.today().year
        self._render_calendar()

    # ── Tooltip ──
    def _tt_show(self, event, d):
        self._tt_hide()
        tw = tk.Toplevel(self)
        tw.wm_overrideredirect(True)
        tw.configure(bg="#2c3e50")
        self._tooltip_win = tw
        tw.wm_geometry(f"+{event.x_root+15}+{event.y_root+10}")

        inner = tk.Frame(tw, bg="#2c3e50", padx=12, pady=8)
        inner.pack()

        dn = ["Montag","Dienstag","Mittwoch","Donnerstag","Freitag","Samstag","Sonntag"]
        tk.Label(inner, text=f"{dn[d.weekday()]}, {d.strftime('%d.%m.%Y')}",
                 bg="#2c3e50", fg="#f39c12", font=("Segoe UI", 10, "bold")).pack(anchor="w")

        sn = self.scraper.get_season(d)
        smap = {"nebensaison": "Nebensaison", "hauptsaison": "Hauptsaison", "basis": "Basispreis"}
        tk.Label(inner, text=smap.get(sn, sn), bg="#2c3e50", fg="#95a5a6",
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(0, 5))

        for room in self.scraper.rooms:
            if room["id"] not in self.active_rooms:
                continue
            av = self.scraper.get_available_count(room["id"], d)
            st = f"{av} verfügbar" if av > 0 else "gesperrt"
            col = "#2ecc71" if av > 0 else "#e74c3c"
            row = tk.Frame(inner, bg="#2c3e50")
            row.pack(fill=tk.X, pady=1)
            tk.Label(row, text=room["name"], bg="#2c3e50", fg="#bdc3c7",
                     font=("Segoe UI", 9)).pack(side=tk.LEFT)
            tk.Label(row, text=st, bg="#2c3e50", fg=col,
                     font=("Segoe UI", 9, "bold")).pack(side=tk.RIGHT, padx=(15, 0))
            if av > 0 and self.scraper.arrangements:
                p = self.scraper.get_price(room["id"], 0, d)
                if p:
                    tk.Label(row, text=f"ab {p:.0f}€", bg="#2c3e50",
                             fg="#bdc3c7", font=("Segoe UI", 8)).pack(side=tk.RIGHT)

    def _tt_hide(self):
        if self._tooltip_win:
            self._tooltip_win.destroy()
            self._tooltip_win = None

    # ── Preisübersicht ──
    def _render_prices(self):
        for w in self.tab_prices.winfo_children():
            w.destroy()

        canvas = tk.Canvas(self.tab_prices, bg=COLORS["white"], highlightthickness=0)
        sb = ttk.Scrollbar(self.tab_prices, orient="vertical", command=canvas.yview)
        sf = tk.Frame(canvas, bg=COLORS["white"])
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=sf, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        if not self.scraper.arrangements:
            tk.Label(sf, text="Keine Arrangements gefunden.", bg=COLORS["white"],
                     fg=COLORS["gray_dark"], font=("Segoe UI", 12)).pack(pady=30)
            return

        # Saisoninfo
        if self.scraper.seasons:
            info = tk.Frame(sf, bg=COLORS["blue_light"], padx=15, pady=10)
            info.pack(fill=tk.X, padx=15, pady=(15, 5))
            tk.Label(info, text="Saisondefinitionen:", bg=COLORS["blue_light"],
                     fg=COLORS["primary"], font=("Segoe UI", 10, "bold")).pack(anchor="w")
            for s in self.scraper.seasons:
                for start, end in s["ranges"]:
                    tk.Label(info,
                             text=f"{s['name'].capitalize()}: "
                                  f"{start.strftime('%d.%m.%Y')} - {end.strftime('%d.%m.%Y')}",
                             bg=COLORS["blue_light"], fg=COLORS["text"],
                             font=("Segoe UI", 9)).pack(anchor="w")

        for arr in self.scraper.arrangements:
            frame = tk.Frame(sf, bg=COLORS["white"], padx=15, pady=10)
            frame.pack(fill=tk.X, padx=15, pady=5)

            tk.Label(frame, text=f"{arr['short_name']} — {arr['name'][:60]}",
                     bg=COLORS["white"], fg=COLORS["primary"],
                     font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 8))

            # Spalten: Kategorie, Zimmer, Basispreis, + je Saison
            season_names = list(arr["season_prices"].keys())
            cols = ["Kategorie", "Zimmer", "Basispreis"] + \
                   [n.capitalize() for n in season_names]

            tbl = tk.Frame(frame, bg="#ddd")
            tbl.pack(fill=tk.X)

            for j, col in enumerate(cols):
                tk.Label(tbl, text=col, bg=COLORS["primary"], fg="white",
                         font=("Segoe UI", 9, "bold"), padx=12, pady=8,
                         anchor="w").grid(row=0, column=j, sticky="nsew", padx=1, pady=1)
                tbl.columnconfigure(j, weight=1 if j == 0 else 0)

            for i, room in enumerate(self.scraper.rooms):
                bg = COLORS["white"] if i % 2 == 0 else COLORS["gray"]
                tk.Label(tbl, text=f"{room['name']} ({room['size']})",
                         bg=bg, fg=COLORS["text"], font=("Segoe UI", 9, "bold"),
                         padx=12, pady=6, anchor="w").grid(
                             row=i+1, column=0, sticky="nsew", padx=1)
                tk.Label(tbl, text=str(room["count"]), bg=bg, fg=COLORS["text"],
                         font=("Segoe UI", 9), padx=12, pady=6).grid(
                             row=i+1, column=1, sticky="nsew", padx=1)

                bp = arr["basis_prices"].get(room["id"])
                tk.Label(tbl, text=f"{bp:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".") if bp else "-",
                         bg=COLORS["green_light"], fg=COLORS["text"],
                         font=("Segoe UI", 9, "bold"), padx=12, pady=6).grid(
                             row=i+1, column=2, sticky="nsew", padx=1)

                for j, sn in enumerate(season_names):
                    sp = arr["season_prices"].get(sn, {}).get(room["id"])
                    txt = f"{sp:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".") if sp else "-"
                    tk.Label(tbl, text=txt, bg=COLORS["blue_light"],
                             fg=COLORS["text"], font=("Segoe UI", 9, "bold"),
                             padx=12, pady=6).grid(
                                 row=i+1, column=3+j, sticky="nsew", padx=1)

    # ── Detailansicht ──
    def _render_details(self):
        for w in self.tab_detail.winfo_children():
            w.destroy()

        canvas = tk.Canvas(self.tab_detail, bg=COLORS["white"], highlightthickness=0)
        sb = ttk.Scrollbar(self.tab_detail, orient="vertical", command=canvas.yview)
        sf = tk.Frame(canvas, bg=COLORS["white"])
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=sf, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        for room in self.scraper.rooms:
            if room["id"] not in self.active_rooms:
                continue

            card = tk.Frame(sf, bg=COLORS["white"], padx=20, pady=15,
                            highlightbackground=room["color"], highlightthickness=2)
            card.pack(fill=tk.X, padx=15, pady=8)

            tk.Label(card, text=f"{room['count']}x {room['name']} ({room['size']})",
                     bg=COLORS["white"], fg=room["color"],
                     font=("Segoe UI", 13, "bold")).pack(anchor="w")

            periods = self._blocked_periods(room["id"])
            if not periods:
                tk.Label(card, text="Keine Sperrdaten — durchgehend verfügbar",
                         bg=COLORS["white"], fg=COLORS["green"],
                         font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=5)
            else:
                tk.Label(card, text=f"{len(periods)} Sperrzeiträume:",
                         bg=COLORS["white"], fg=COLORS["red"],
                         font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(5, 2))
                for s, e in periods:
                    days = (e - s).days + 1
                    tk.Label(card,
                             text=f"  {s.strftime('%d.%m.%Y')} — {e.strftime('%d.%m.%Y')} ({days} Tage)",
                             bg=COLORS["white"], fg=COLORS["text"],
                             font=("Segoe UI", 9)).pack(anchor="w")

            if self.scraper.arrangements:
                pf = tk.Frame(card, bg=COLORS["white"])
                pf.pack(anchor="w", pady=(10, 0))
                for arr in self.scraper.arrangements:
                    bp = arr["basis_prices"].get(room["id"])
                    parts = [f"Basis: {bp:.0f}€"] if bp else []
                    for sn, sp in arr["season_prices"].items():
                        p = sp.get(room["id"])
                        if p:
                            parts.append(f"{sn.capitalize()}: {p:.0f}€")
                    if parts:
                        tk.Label(pf, text=f"{arr['short_name']}: {' | '.join(parts)}",
                                 bg=COLORS["white"], fg=COLORS["text"],
                                 font=("Segoe UI", 9)).pack(anchor="w")

    def _blocked_periods(self, room_id):
        blocked = self.scraper.blocked.get(room_id, set())
        if not blocked:
            return []
        dates = sorted(blocked)
        periods = []
        start = prev = None
        for dk in dates:
            d = date.fromisoformat(dk)
            if start is None:
                start = prev = d
            elif (d - prev).days == 1:
                prev = d
            else:
                periods.append((start, prev))
                start = prev = d
        if start:
            periods.append((start, prev))
        return periods

    # ── Excel Export ──
    def _export_excel(self):
        if not self.scraper.rooms:
            messagebox.showwarning("Keine Daten", "Bitte zuerst Daten laden.")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror("Fehler", "openpyxl nicht installiert.\npip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"Hotel_Aquarius_{self.current_year}.xlsx")
        if not path:
            return

        self.status_var.set("Exportiere...")
        if not self.progress.winfo_ismapped():
            self.progress.pack(fill=tk.X, side=tk.BOTTOM)
        self.progress.start()

        def worker():
            try:
                self._write_excel(path)
                self.after(0, lambda: self._export_ok(path))
            except Exception as e:
                self.after(0, lambda: self._export_err(str(e)))

        threading.Thread(target=worker, daemon=True).start()

    def _write_excel(self, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Verfügbarkeit"

        hfill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
        hfont = Font(color="FFFFFF", bold=True, size=10)
        gfill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        rfill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
        bfill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        thin = Border(left=Side(style='thin', color='DDDDDD'),
                      right=Side(style='thin', color='DDDDDD'),
                      top=Side(style='thin', color='DDDDDD'),
                      bottom=Side(style='thin', color='DDDDDD'))

        today = date.today()
        end_d = date(self.current_year, 12, 31)
        dates = []
        d = today
        while d <= end_d:
            dates.append(d)
            d += timedelta(days=1)

        ws.cell(row=1, column=1, value="Kategorie").font = hfont
        ws.cell(row=1, column=1).fill = hfill
        ws.column_dimensions['A'].width = 30

        for i, d in enumerate(dates):
            c = ws.cell(row=1, column=i+2, value=d.strftime("%d.%m"))
            c.font = hfont; c.fill = hfill
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(i+2)].width = 7

        row = 2
        for room in self.scraper.rooms:
            ws.cell(row=row, column=1,
                    value=f"{room['count']}x {room['name']}").font = Font(bold=True, size=10)
            ws.cell(row=row, column=1).fill = bfill

            for i, d in enumerate(dates):
                av = self.scraper.get_available_count(room["id"], d)
                c = ws.cell(row=row, column=i+2, value=av)
                c.alignment = Alignment(horizontal="center")
                c.border = thin
                if av > 0:
                    c.fill = gfill; c.font = Font(color="1E8449", bold=True)
                else:
                    c.fill = rfill; c.font = Font(color="C0392B", bold=True)
            row += 1

            for arr in self.scraper.arrangements:
                ws.cell(row=row, column=1,
                        value=f"  {arr['short_name']} p.P.").font = Font(size=9, italic=True)
                for i, d in enumerate(dates):
                    p = self.scraper.get_price(room["id"],
                                                self.scraper.arrangements.index(arr), d)
                    if p:
                        c = ws.cell(row=row, column=i+2, value=p)
                        c.number_format = '#,##0.00 €'
                        c.alignment = Alignment(horizontal="center")
                        c.font = Font(size=8); c.border = thin
                        season = self.scraper.get_season(d)
                        c.fill = bfill if season != "basis" else gfill
                row += 1
            row += 1

        ws.freeze_panes = "B2"

        # Preise-Sheet
        ws2 = wb.create_sheet("Preise")
        ws2.cell(row=1, column=1, value="Preisübersicht Hotel Aquarius").font = Font(size=14, bold=True)
        r = 3
        for arr in self.scraper.arrangements:
            ws2.cell(row=r, column=1, value=f"{arr['short_name']} — {arr['name'][:50]}").font = Font(size=11, bold=True)
            r += 1
            skeys = list(arr["season_prices"].keys())
            cols = ["Kategorie", "Zimmer", "Basispreis"] + [k.capitalize() for k in skeys]
            for j, col in enumerate(cols):
                c = ws2.cell(row=r, column=j+1, value=col)
                c.font = hfont; c.fill = hfill
            r += 1
            for room in self.scraper.rooms:
                ws2.cell(row=r, column=1, value=f"{room['name']} ({room['size']})").font = Font(bold=True)
                ws2.cell(row=r, column=2, value=room["count"])
                bp = arr["basis_prices"].get(room["id"])
                if bp:
                    c = ws2.cell(row=r, column=3, value=bp)
                    c.number_format = '#,##0.00 €'; c.font = Font(bold=True)
                for j, sn in enumerate(skeys):
                    sp = arr["season_prices"].get(sn, {}).get(room["id"])
                    if sp:
                        c = ws2.cell(row=r, column=4+j, value=sp)
                        c.number_format = '#,##0.00 €'; c.font = Font(bold=True)
                r += 1
            r += 1

        for letter in 'ABCDEF':
            ws2.column_dimensions[letter].width = 25
        wb.save(path)

    def _export_ok(self, path):
        self.progress.stop(); self.progress.pack_forget()
        self.status_var.set(f"Excel exportiert: {path}")
        messagebox.showinfo("Export", f"Gespeichert:\n{path}")

    def _export_err(self, err):
        self.progress.stop(); self.progress.pack_forget()
        self.status_var.set(f"Export-Fehler: {err}")
        messagebox.showerror("Fehler", f"Export fehlgeschlagen:\n{err}")


# ── Start ──
if __name__ == "__main__":
    if requests is None:
        import sys
        print("=" * 50)
        print("Benötigte Pakete nicht installiert!")
        print("pip install requests beautifulsoup4 openpyxl")
        print("=" * 50)
        sys.exit(1)
    HotelApp().mainloop()
