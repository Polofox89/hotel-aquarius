#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tagesbuffet-Generator – Hotel Aquarius
=======================================
Erstellt Buffet-Menübilder (1080×1920 px, Hochformat) mit
automatischer Excel-Archivierung.

Verwendung:
    python tagesbuffet_generator.py            # Interaktiver Modus
    python tagesbuffet_generator.py --voice    # Spracheingabe (optional)
    python tagesbuffet_generator.py --output ./output --logo ./logo.jpg
"""

import os
import sys
import argparse
from datetime import datetime
from pathlib import Path
from typing import Optional

# Windows-Terminal: UTF-8 erzwingen (verhindert UnicodeEncodeError mit ä/ö/ü/ß/═)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stdin, "reconfigure"):
    sys.stdin.reconfigure(encoding="utf-8")

# ── Pflicht-Abhängigkeiten ────────────────────────────────────────────────────
try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:
    print("Fehler: Pillow nicht installiert.")
    print("       pip install Pillow")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font as XLFont, Alignment
except ImportError:
    print("Fehler: openpyxl nicht installiert.")
    print("       pip install openpyxl")
    sys.exit(1)

# ── Optionale Spracherkennung ─────────────────────────────────────────────────
SPEECH_AVAILABLE = False
try:
    import speech_recognition as sr
    SPEECH_AVAILABLE = True
except ImportError:
    pass  # Spracheingabe deaktiviert, Fallback auf Tastatur


# ══════════════════════════════════════════════════════════════════════════════
# Konfiguration
# ══════════════════════════════════════════════════════════════════════════════

# Pfade (überschreibbar per Umgebungsvariable oder CLI-Argument)
def _find_logo() -> Path:
    """Sucht das Hotel-Logo an mehreren üblichen Orten."""
    env = os.getenv("BUFFET_LOGO")
    candidates = []
    if env:
        candidates.append(Path(env))
    script_dir = Path(__file__).parent
    candidates += [
        script_dir / "Logo_JPG.jpg",
        script_dir / "logo.png",
        script_dir / "logo.jpg",
        Path("C:/hotel/Logo_JPG.jpg"),
    ]
    for c in candidates:
        try:
            if c.exists():
                return c
        except OSError:
            continue
    # Fallback (existiert evtl. nicht – Generator prüft das)
    return script_dir / "Logo_JPG.jpg"

LOGO_PATH   = _find_logo()
OUTPUT_DIR  = Path(os.getenv("BUFFET_OUTPUT", str(Path(__file__).parent / "buffet")))
ARCHIV_FILE = OUTPUT_DIR / "Tagesbuffet_Archiv.xlsx"

# Bildabmessungen
IMG_W = 1080
IMG_H = 1920

# Farben (RGB)
COLOR_BG    = (255, 255, 255)   # Weißer Hintergrund
COLOR_TEXT  = ( 44,  44,  44)   # #2c2c2c – Haupttext & Speisen
COLOR_DATE  = (102, 102, 102)   # #666666 – Datum
COLOR_LABEL = (153, 153, 153)   # #999999 – Kategorie-Label
COLOR_LINE  = (204, 204, 204)   # #cccccc – Trennlinien

# Excel-Farben (Hex ohne #)
XL_HEADER_HEX  = "0d4f6b"   # Dunkelblau Header
XL_ROW_ALT_HEX = "e8f4f8"   # Hellblau Alternativzeile

# Schriftgrößen: pt → px (96 dpi: px = pt × 96/72)
def _pt(points: float) -> int:
    return round(points * 96 / 72)

FONT_SIZE_TITLE = _pt(55)   # 73 px – "TAGESBUFFET"
FONT_SIZE_DATE  = _pt(28)   # 37 px – Datum
FONT_SIZE_LABEL = _pt(22)   # 29 px – Kategorie-Label
FONT_SIZE_DISH  = _pt(50)   # 67 px – Speisen

# Layout-Abstände (px)
LOGO_W           = 675
SPACING_LOGO     = 50    # Abstand Logo → Titel
SPACING_TITLE    = 18    # Abstand Titel → Datum
SPACING_DATE     = 55    # Abstand Datum → erste Linie
SPACING_PRE_LINE = 28    # Abstand vor Trennlinie
SPACING_POST_LINE= 18    # Abstand Linie → Label
SPACING_POST_LABEL= 32   # Abstand Label → erste Speise
SPACING_DISH_GAP = 22    # Abstand zwischen Speisen einer Kategorie
SPACING_POST_CAT = 18    # Abstand nach letzter Speise
LINE_THICKNESS   = 1
LINE_MARGIN      = 60    # Horizontaler Einzug der Linie

# Deutsche Lokalisierung
WEEKDAYS_DE = [
    "Montag", "Dienstag", "Mittwoch", "Donnerstag",
    "Freitag", "Samstag", "Sonntag"
]
MONTHS_DE = [
    "", "Januar", "Februar", "März", "April", "Mai", "Juni",
    "Juli", "August", "September", "Oktober", "November", "Dezember"
]

# Menü-Kategorien: (key, Anzeigename, min, max, Standardwerte)
CATEGORIES = [
    ("suppe",       "SUPPE",         1, 1, []),
    ("haupt",       "HAUPTGERICHTE", 2, 2, []),
    ("pasta",       "PASTA",         0, 1, []),
    ("beilagen",    "BEILAGEN",      3, 4, []),
    ("salate",      "SALATE",        1, 1, ["Salatbuffet"]),
    ("partypfanne", "PARTYPFANNE",   1, 1, []),
    ("dessert",     "DESSERT",       1, 1, ["Dessertbuffet"]),
]


# ══════════════════════════════════════════════════════════════════════════════
# Generator-Klasse
# ══════════════════════════════════════════════════════════════════════════════

class TagesbuffetGenerator:
    """
    Erstellt Tagesbuffet-Menübilder und verwaltet das Excel-Archiv.

    Typische Nutzung:
        gen = TagesbuffetGenerator()
        date, menu = gen.get_menu_interactive()
        gen.create_image(date, menu)
        gen.update_excel(date, menu)
    """

    def __init__(
        self,
        image_dir: Optional[Path] = None,
        archiv_dir: Optional[Path] = None,
    ) -> None:
        self.image_dir: Path = Path(image_dir) if image_dir else OUTPUT_DIR
        self.archiv_dir: Path = Path(archiv_dir) if archiv_dir else OUTPUT_DIR
        self.fonts: dict = {}
        self._load_fonts()

    # ── Pfad-Verwaltung ───────────────────────────────────────────────────────

    @property
    def archiv_file(self) -> Path:
        """Pfad zur Excel-Archivdatei (im konfigurierten Archiv-Ordner)."""
        return self.archiv_dir / "Tagesbuffet_Archiv.xlsx"

    def set_image_dir(self, path) -> None:
        """Setzt den Ausgabeordner für die JPG-Bilder."""
        self.image_dir = Path(path)

    def set_archiv_dir(self, path) -> None:
        """Setzt den Ordner für Excel-Archiv und Konfigurations-Dateien."""
        self.archiv_dir = Path(path)

    # ── Font-Verwaltung ───────────────────────────────────────────────────────

    def _load_fonts(self) -> None:
        """
        Lädt TrueType-Schriften aus systemüblichen Pfaden.
        Fetter Schnitt für Titel/Speisen, Regular für Datum/Label.
        Fallback auf PIL-Standardschrift wenn keine gefunden wird.
        """
        bold_candidates = [
            # Windows – elegant
            "C:/Windows/Fonts/palab.ttf",       # Palatino Linotype Bold
            "C:/Windows/Fonts/cambriab.ttf",    # Cambria Bold
            "C:/Windows/Fonts/georgiab.ttf",    # Georgia Bold
            "C:/Windows/Fonts/Candarab.ttf",    # Candara Bold
            "C:/Windows/Fonts/calibrib.ttf",    # Calibri Bold (Fallback)
            "C:/Windows/Fonts/arialbd.ttf",     # Arial Bold (Fallback)
            # Linux
            "/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSerif-Bold.ttf",
            # macOS
            "/Library/Fonts/Palatino.ttc",
            "/System/Library/Fonts/Helvetica.ttc",
        ]
        regular_candidates = [
            # Windows – elegant
            "C:/Windows/Fonts/pala.ttf",        # Palatino Linotype
            "C:/Windows/Fonts/cambria.ttc",     # Cambria
            "C:/Windows/Fonts/georgia.ttf",     # Georgia
            "C:/Windows/Fonts/Candara.ttf",     # Candara
            "C:/Windows/Fonts/calibri.ttf",     # Calibri (Fallback)
            "C:/Windows/Fonts/arial.ttf",       # Arial (Fallback)
            # Linux
            "/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSerif-Regular.ttf",
            # macOS
            "/Library/Fonts/Palatino.ttc",
            "/System/Library/Fonts/Helvetica.ttc",
        ]

        def try_load(candidates: list, size: int) -> ImageFont.FreeTypeFont:
            for path in candidates:
                if Path(path).exists():
                    try:
                        return ImageFont.truetype(path, size)
                    except Exception:
                        continue
            # Fallback – Pillow-Standardschrift (keine Größenkontrolle)
            print(f"  Warnung: Keine TrueType-Schrift gefunden (Größe {size}px). "
                  "Verwende Standardschrift.")
            return ImageFont.load_default()

        self.fonts = {
            "title": try_load(bold_candidates,    FONT_SIZE_TITLE),
            "date":  try_load(regular_candidates, FONT_SIZE_DATE),
            "label": try_load(regular_candidates, FONT_SIZE_LABEL),
            "dish":  try_load(bold_candidates,    FONT_SIZE_DISH),
        }

    # ── Hilfsmethoden ─────────────────────────────────────────────────────────

    @staticmethod
    def format_date_de(date: datetime) -> str:
        """Gibt das Datum auf Deutsch zurück: 'Dienstag, 03. März 2026'."""
        wochentag = WEEKDAYS_DE[date.weekday()]
        monat     = MONTHS_DE[date.month]
        return f"{wochentag}, {date.day:02d}. {monat} {date.year}"

    def _text_wh(self, text: str, font) -> tuple[int, int]:
        """Gibt (Breite, Höhe) eines Textes zurück."""
        dummy = Image.new("RGB", (1, 1))
        bbox  = ImageDraw.Draw(dummy).textbbox((0, 0), text, font=font)
        return bbox[2] - bbox[0], bbox[3] - bbox[1]

    def _draw_centered(
        self,
        draw: ImageDraw.Draw,
        y: int,
        text: str,
        font,
        color: tuple,
    ) -> int:
        """
        Zeichnet Text horizontal auf IMG_W zentriert.
        Gibt die Texthöhe zurück.
        """
        w, h = self._text_wh(text, font)
        x = (IMG_W - w) // 2
        draw.text((x, y), text, font=font, fill=color)
        return h

    def _draw_separator(self, draw: ImageDraw.Draw, y: int) -> None:
        """Zeichnet eine horizontale Trennlinie."""
        draw.line(
            [(LINE_MARGIN, y), (IMG_W - LINE_MARGIN, y)],
            fill=COLOR_LINE,
            width=LINE_THICKNESS,
        )

    def _calc_content_height(self, logo_h: int, menu: dict, sp: dict = None) -> int:
        """
        Berechnet die Gesamthöhe aller Inhaltselemente.
        Wird für das vertikale Zentrieren / die adaptive Skalierung benötigt.

        Args:
            sp: Optional dict mit skalierten Abstandswerten. Falls None,
                werden die Modul-Konstanten verwendet.
        """
        if sp is None:
            sp = {
                "logo": SPACING_LOGO, "title": SPACING_TITLE, "date": SPACING_DATE,
                "pre_line": SPACING_PRE_LINE, "post_line": SPACING_POST_LINE,
                "post_label": SPACING_POST_LABEL, "dish_gap": SPACING_DISH_GAP,
                "post_cat": SPACING_POST_CAT,
            }

        h = logo_h + sp["logo"]

        _, th = self._text_wh("TAGESBUFFET", self.fonts["title"])
        h += th + sp["title"]

        _, dh = self._text_wh("Mustertext", self.fonts["date"])
        h += dh + sp["date"]

        for key, label, *_ in CATEGORIES:
            dishes = menu.get(key, [])
            if not dishes:
                continue

            h += sp["pre_line"] + LINE_THICKNESS + sp["post_line"]

            _, lh = self._text_wh(label, self.fonts["label"])
            h += lh + sp["post_label"]

            for dish in dishes:
                _, dih = self._text_wh(dish, self.fonts["dish"])
                h += dih + sp["dish_gap"]

            h += sp["post_cat"]

        return h

    # Mindestabstände – darunter werden Label und Speisen visuell unleserlich
    _SPACING_FLOOR = {
        "logo": 28, "title": 10, "date": 30,
        "pre_line": 14, "post_line": 10,
        "post_label": 16, "dish_gap": 12, "post_cat": 10,
    }
    # Bevorzugte Abstände (großzügig, gut lesbar)
    _SPACING_FULL = None  # wird unten dynamisch befüllt

    @classmethod
    def _interp_spacings(cls, s: float) -> dict:
        """Interpoliert Abstände zwischen Floor (s=0) und Full (s=1)."""
        full = {
            "logo": SPACING_LOGO, "title": SPACING_TITLE, "date": SPACING_DATE,
            "pre_line": SPACING_PRE_LINE, "post_line": SPACING_POST_LINE,
            "post_label": SPACING_POST_LABEL, "dish_gap": SPACING_DISH_GAP,
            "post_cat": SPACING_POST_CAT,
        }
        floor = cls._SPACING_FLOOR
        s = max(0.0, min(1.0, s))
        return {k: int(round(floor[k] + s * (full[k] - floor[k]))) for k in full}

    def _fit_spacings(self, logo_h: int, menu: dict, available_h: int) -> dict:
        """
        Findet den größtmöglichen Skalierungsfaktor s ∈ [0..1], bei dem
        der Inhalt vollständig in `available_h` Pixel passt.
        Großzügiger Abstand (s=1), wenn der Inhalt klein genug ist,
        ansonsten reduziert er sich Richtung _SPACING_FLOOR.
        """
        # 21 Stufen von s=1.0 (großzügig) bis s=0.0 (eng) durchprobieren
        for step in range(20, -1, -1):
            s = step / 20.0
            sp = self._interp_spacings(s)
            if self._calc_content_height(logo_h, menu, sp) <= available_h:
                return sp
        # Worst case: kleinster Abstand
        return self._interp_spacings(0.0)

    # ── Bildgenerierung ───────────────────────────────────────────────────────

    def render_image(self, date: datetime, menu: dict) -> Image.Image:
        """
        Rendert das Tagesbuffet-Bild (1080×1920 px) ohne es zu speichern.
        Dient der Live-Vorschau in der GUI.

        Args:
            date: Datum des Buffets.
            menu: Dict mit Kategorie-Schlüsseln → Listen von Speisen.

        Returns:
            PIL-Image-Objekt (RGB).
        """
        img  = Image.new("RGB", (IMG_W, IMG_H), color=COLOR_BG)
        draw = ImageDraw.Draw(img)

        # ── Logo laden ───────────────────────────────────────────────────────
        logo_h = 0
        logo_img: Optional[Image.Image] = None

        if LOGO_PATH.exists():
            try:
                raw_logo = Image.open(LOGO_PATH).convert("RGBA")
                ratio    = LOGO_W / raw_logo.width
                logo_h   = round(raw_logo.height * ratio)
                raw_logo = raw_logo.resize((LOGO_W, logo_h), Image.LANCZOS)

                logo_bg = Image.new("RGB", (LOGO_W, logo_h), COLOR_BG)
                if raw_logo.mode == "RGBA":
                    logo_bg.paste(raw_logo, mask=raw_logo.split()[3])
                else:
                    logo_bg.paste(raw_logo)
                logo_img = logo_bg
            except Exception as e:
                print(f"  Warnung: Logo konnte nicht geladen werden – {e}")

        # ── Adaptive Abstände: bei vollem Menü wird automatisch gestaucht ────
        TOP_MARGIN = 80
        available_h = IMG_H - 2 * TOP_MARGIN
        sp = self._fit_spacings(logo_h, menu, available_h)

        # ── Startposition: Inhalt vertikal zentrieren ─────────────────────────
        content_h = self._calc_content_height(logo_h, menu, sp)
        y = max(TOP_MARGIN, (IMG_H - content_h) // 2)

        # ── Logo einfügen ─────────────────────────────────────────────────────
        if logo_img is not None:
            logo_x = (IMG_W - LOGO_W) // 2
            img.paste(logo_img, (logo_x, y))
            y += logo_h

        y += sp["logo"]

        # ── Titel ─────────────────────────────────────────────────────────────
        h  = self._draw_centered(draw, y, "TAGESBUFFET", self.fonts["title"], COLOR_TEXT)
        y += h + sp["title"]

        # ── Datum (deutsch) ───────────────────────────────────────────────────
        h  = self._draw_centered(draw, y, self.format_date_de(date), self.fonts["date"], COLOR_DATE)
        y += h + sp["date"]

        # ── Kategorien ────────────────────────────────────────────────────────
        for key, label, *_ in CATEGORIES:
            dishes = menu.get(key, [])
            if not dishes:
                continue

            y += sp["pre_line"]
            self._draw_separator(draw, y)
            y += LINE_THICKNESS + sp["post_line"]

            h  = self._draw_centered(draw, y, label, self.fonts["label"], COLOR_LABEL)
            y += h + sp["post_label"]

            for dish in dishes:
                h  = self._draw_centered(draw, y, dish, self.fonts["dish"], COLOR_TEXT)
                y += h + sp["dish_gap"]

            y += sp["post_cat"]

        return img

    def create_image(self, date: datetime, menu: dict) -> Path:
        """
        Erstellt das Tagesbuffet-Bild und speichert es als JPEG.

        Args:
            date: Datum des Buffets.
            menu: Dict mit Kategorie-Schlüsseln → Listen von Speisen.

        Returns:
            Pfad zur gespeicherten JPEG-Datei.

        Raises:
            OSError: Bei Schreibfehler in das Ausgabeverzeichnis.
        """
        self.image_dir.mkdir(parents=True, exist_ok=True)

        img = self.render_image(date, menu)

        filename    = f"tagesbuffet_{date.strftime('%d_%m_%Y')}.jpg"
        output_path = self.image_dir / filename
        img.save(output_path, "JPEG", quality=95, optimize=True)
        print(f"  Bild gespeichert:      {output_path}")
        return output_path

    # ── Excel-Archivierung ────────────────────────────────────────────────────

    def update_excel(self, date: datetime, menu: dict) -> None:
        """
        Fügt das neue Menü dem Excel-Archiv hinzu.
        Erstellt die Datei neu falls sie noch nicht existiert.

        Args:
            date: Datum des Buffets.
            menu: Dict mit Kategorie-Schlüsseln → Listen von Speisen.
        """
        self.archiv_dir.mkdir(parents=True, exist_ok=True)
        archiv_file = self.archiv_file

        # Workbook laden oder neu anlegen
        if archiv_file.exists():
            try:
                wb = load_workbook(archiv_file)
                ws = wb.active
            except Exception as e:
                print(f"  Warnung: Archiv-Datei beschädigt ({e}), erstelle neu.")
                wb = Workbook()
                ws = wb.active
                ws.title = "Tagesbuffet-Archiv"
                self._write_excel_header(ws)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Tagesbuffet-Archiv"
            self._write_excel_header(ws)

        # Nächste freie Zeile
        next_row  = ws.max_row + 1
        datum_str = date.strftime("%d.%m.%Y")

        for key, label, *_ in CATEGORIES:
            for dish in menu.get(key, []):
                self._write_excel_row(ws, next_row, datum_str, label, dish)
                next_row += 1

        # Spaltenbreiten (Datum | Kategorie | Speise)
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 42

        wb.save(archiv_file)
        print(f"  Excel-Archiv aktualisiert: {archiv_file}")

    def _write_excel_header(self, ws) -> None:
        """Schreibt die Kopfzeile mit dunkelblauer Formatierung."""
        header_fill  = PatternFill("solid", fgColor=XL_HEADER_HEX)
        header_font  = XLFont(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")

        for col, title in enumerate(["Datum", "Kategorie", "Speise"], start=1):
            cell            = ws.cell(row=1, column=col, value=title)
            cell.fill       = header_fill
            cell.font       = header_font
            cell.alignment  = header_align

        ws.row_dimensions[1].height = 22

    def _write_excel_row(
        self, ws, row: int, datum: str, kategorie: str, speise: str
    ) -> None:
        """
        Schreibt eine Datenzeile mit alternierender Zeilenfarbe.
        Gerade Zeilen: hellblau, ungerade: weiß (Header-Zeile ausgenommen).
        """
        alt_fill  = PatternFill("solid", fgColor=XL_ROW_ALT_HEX)
        row_align = Alignment(vertical="center")
        use_alt   = row % 2 == 0   # Jede zweite Datenzeile einfärben

        for col, value in enumerate([datum, kategorie, speise], start=1):
            cell           = ws.cell(row=row, column=col, value=value)
            cell.alignment = row_align
            if use_alt:
                cell.fill = alt_fill

    # ── Interaktive Eingabe ───────────────────────────────────────────────────

    def _input_date(self) -> datetime:
        """Fragt das Datum ab. Leere Eingabe → heutiges Datum."""
        print()
        raw = input("  Datum [TT.MM.JJJJ, Enter = heute]: ").strip()
        if not raw:
            return datetime.today()
        try:
            return datetime.strptime(raw, "%d.%m.%Y")
        except ValueError:
            print("  Ungültiges Format – verwende heutiges Datum.")
            return datetime.today()

    def _input_dishes(
        self,
        label: str,
        min_d: int,
        max_d: int,
        default: list[str],
        hint: str = "",
    ) -> list[str]:
        """
        Fragt die Gerichte einer Kategorie interaktiv ab.

        Args:
            label:   Anzeigename der Kategorie.
            min_d:   Mindestanzahl Gerichte (Pflicht).
            max_d:   Maximalanzahl Gerichte.
            default: Vorausgefüllte Werte (Enter übernimmt diese).
            hint:    Optionaler Hinweis (z.B. Reihenfolge-Tipp).

        Returns:
            Liste der eingegebenen Gerichte.
        """
        print(f"\n  ── {label}  ({min_d}–{max_d} Gerichte) ──")
        if hint:
            print(f"  {hint}")
        if default:
            print(f"  Standard: {', '.join(default)}")

        dishes: list[str] = []

        for i in range(1, max_d + 1):
            def_val  = default[i - 1] if i <= len(default) else ""
            optional = i > min_d

            if def_val:
                prompt = f"  {i}. [Enter = '{def_val}']: "
            elif optional:
                prompt = f"  {i}. [Enter = fertig]:    "
            else:
                prompt = f"  {i}. (Pflicht):            "

            while True:
                raw = input(prompt).strip()

                if raw:
                    dishes.append(raw)
                    break
                elif def_val:
                    dishes.append(def_val)
                    break
                elif optional:
                    return dishes
                else:
                    print(f"  Mindestens {min_d} Gericht(e) erforderlich.")

        return dishes

    def get_menu_interactive(self) -> tuple[datetime, dict]:
        """
        Interaktive Menü-Eingabe via Terminal.

        Returns:
            Tuple (date, menu) – Datum und Dict mit Kategorien → Speisen.
        """
        print("\n" + "═" * 52)
        print("   TAGESBUFFET-GENERATOR  –  Hotel Aquarius")
        print("═" * 52)

        date = self._input_date()
        print(f"\n  Datum: {self.format_date_de(date)}")

        hints = {
            "beilagen": "Tipp: Gemüse zuerst, dann Stärkebeilagen eingeben.",
        }

        menu: dict = {}
        for key, label, min_d, max_d, default in CATEGORIES:
            menu[key] = self._input_dishes(
                label, min_d, max_d, default, hint=hints.get(key, "")
            )

        return date, menu

    # ── Spracheingabe (optional) ──────────────────────────────────────────────

    def _listen(self, recognizer, mic, prompt: str) -> str:
        """
        Nimmt eine Spracheingabe auf und gibt den transkribierten Text zurück.
        Gibt einen leeren String zurück wenn nichts erkannt wurde.
        """
        print(f"  [Mikro] {prompt}")
        try:
            with mic as source:
                recognizer.adjust_for_ambient_noise(source, duration=0.5)
                audio = recognizer.listen(source, timeout=8, phrase_time_limit=6)
            text = recognizer.recognize_google(audio, language="de-DE")
            print(f"  Erkannt: \"{text}\"")
            return text
        except sr.WaitTimeoutError:
            print("  Zeitüberschreitung – kein Ton erkannt.")
            return ""
        except sr.UnknownValueError:
            print("  Sprache nicht verstanden.")
            return ""
        except sr.RequestError as e:
            print(f"  Erkennungs-Fehler: {e}")
            return ""

    def get_menu_voice(self) -> tuple[datetime, dict]:
        """
        Menü-Eingabe via Sprache.
        Fallback auf Tastatur wenn speech_recognition nicht verfügbar.

        Returns:
            Tuple (date, menu) – Datum und Dict mit Kategorien → Speisen.
        """
        if not SPEECH_AVAILABLE:
            print("  Spracheingabe nicht verfügbar.")
            print("  Installation: pip install SpeechRecognition pyaudio")
            print("  Wechsle zu Tastatur-Eingabe ...\n")
            return self.get_menu_interactive()

        r   = sr.Recognizer()
        mic = sr.Microphone()

        print("\n" + "═" * 52)
        print("   TAGESBUFFET-GENERATOR  –  Spracheingabe")
        print("═" * 52)
        print("  Datum: heutiges Datum wird verwendet.")

        date = datetime.today()
        menu: dict = {}

        for key, label, min_d, max_d, default in CATEGORIES:
            if default:
                # Standardwerte direkt übernehmen (z.B. Salatbuffet, Dessertbuffet)
                menu[key] = list(default)
                print(f"\n  {label}: {', '.join(default)} (Standard)")
                continue

            print(f"\n  ── {label}  ({min_d}–{max_d} Gerichte) ──")
            dishes: list[str] = []

            for i in range(1, max_d + 1):
                text = self._listen(r, mic, f"Gericht {i} sprechen (oder Enter drücken):")

                if not text:
                    raw = input("  Tastatur-Eingabe (Enter = überspringen): ").strip()
                    text = raw

                if text:
                    dishes.append(text.capitalize())
                elif i > min_d:
                    break

            menu[key] = dishes if dishes else list(default)

        return date, menu

    # ── Bestätigung ───────────────────────────────────────────────────────────

    def confirm_menu(self, date: datetime, menu: dict) -> bool:
        """
        Zeigt das Menü zur Bestätigung an.

        Returns:
            True wenn der Benutzer bestätigt, sonst False.
        """
        print("\n" + "─" * 52)
        print(f"  {self.format_date_de(date)}")
        print("─" * 52)

        for key, label, *_ in CATEGORIES:
            dishes = menu.get(key, [])
            if dishes:
                print(f"  {label}:")
                for dish in dishes:
                    print(f"    • {dish}")

        print("─" * 52)
        answer = input("  Menü erstellen? [J/n]: ").strip().lower()
        return answer in ("", "j", "ja", "y", "yes")

    # ── Hauptmethode ──────────────────────────────────────────────────────────

    def run(self, voice: bool = False) -> None:
        """
        Hauptablauf: Eingabe → Bestätigung → Bild erstellen → Excel aktualisieren.

        Args:
            voice: Wenn True, wird Spracheingabe verwendet.
        """
        try:
            if voice:
                date, menu = self.get_menu_voice()
            else:
                date, menu = self.get_menu_interactive()

            if not self.confirm_menu(date, menu):
                print("\n  Abgebrochen. Auf Wiedersehen.")
                return

            print("\n  Erstelle Menü-Bild ...")
            img_path = self.create_image(date, menu)

            print("  Aktualisiere Excel-Archiv ...")
            self.update_excel(date, menu)

            print(f"\n  ✓ Fertig!")
            print(f"    Bild:   {img_path}")
            print(f"    Archiv: {self.archiv_file}\n")

        except KeyboardInterrupt:
            print("\n\n  Abgebrochen.")
        except Exception as e:
            print(f"\n  Fehler: {e}")
            raise


# ══════════════════════════════════════════════════════════════════════════════
# Einstiegspunkt
# ══════════════════════════════════════════════════════════════════════════════

def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Tagesbuffet-Generator – Hotel Aquarius",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Beispiele:
  python tagesbuffet_generator.py
  python tagesbuffet_generator.py --voice
  python tagesbuffet_generator.py --output ./output --logo ./Logo.jpg
        """,
    )
    parser.add_argument(
        "--voice",
        action="store_true",
        help="Spracheingabe aktivieren (benötigt SpeechRecognition + pyaudio)",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        metavar="PFAD",
        help="Ausgabeordner (Standard: /mnt/user-data/outputs)",
    )
    parser.add_argument(
        "--logo",
        type=str,
        default=None,
        metavar="DATEI",
        help="Logo-Pfad (Standard: C:/hotel/Logo_JPG.jpg)",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()

    # Pfade überschreiben (CLI hat Vorrang vor Umgebungsvariablen)
    global LOGO_PATH
    if args.logo:
        LOGO_PATH = Path(args.logo)

    out = Path(args.output) if args.output else None
    generator = TagesbuffetGenerator(image_dir=out, archiv_dir=out)
    generator.run(voice=args.voice)


if __name__ == "__main__":
    main()
